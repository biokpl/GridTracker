#!/usr/bin/env python3
"""
GridTracker Server - port 5050
- ATR_Sonuc.xlsx ve Destek_Direc_Seviyeleri.xlsx okur
- GET /api/stock/{SYM}  → hisse verisi (fiyat + ATR + destek/direnç)
- GET /api/all          → tüm hisseler
- GET /api/health       → durum
- POST /api/notify      → push bildirimi gönder
- GET /                 → bist_tracker.html
- Arka planda 60s'de bir Firebase'e push eder
Stdlib only + openpyxl (otomatik install edilir)
"""
import subprocess, sys

for _pkg in ['openpyxl', 'pywebpush']:
    try:
        __import__(_pkg.replace('-', '_'))
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', _pkg, '-q'],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

from openpyxl import load_workbook
from http.server import HTTPServer, ThreadingHTTPServer, SimpleHTTPRequestHandler
import urllib.request, json, socket, threading, time, os, logging

FIREBASE_URL = 'https://grid-tracker-73ed2-default-rtdb.europe-west1.firebasedatabase.app'
NTFY_TOPIC   = 'GridTracker-bkpl-07'
NTFY_TOPIC_ALERT = NTFY_TOPIC + '-acil'   # kritik (ÇIK) → alarm sesli topic
PORT = 5050
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ATR_FILE   = os.path.join(BASE_DIR, 'ATR_Sonuc.xlsx')
DD_FILE    = os.path.join(BASE_DIR, 'Destek_Direc_Seviyeleri.xlsx')
VAPID_FILE = os.path.join(BASE_DIR, 'vapid_keys.json')
LOG_FILE   = os.path.join(BASE_DIR, 'server.log')

VAPID_CLAIMS = {'sub': 'mailto:admin@gridtracker.local'}

# ── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.FileHandler(LOG_FILE, encoding='utf-8')]
)
slog = logging.getLogger('server')

# Bellekte tutulan veri
_stocks = {}
_stocks_ts = 0
_lock = threading.Lock()

# ---------------------------------------------------------------------------
# Excel okuma
# ---------------------------------------------------------------------------

def _val(cell):
    v = cell.value
    if v is None: return None
    try: return float(v)
    except: return str(v).strip()

def read_excel():
    atr = {}
    dd  = {}

    # ATR_Sonuc.xlsx
    if os.path.exists(ATR_FILE):
        try:
            wb = load_workbook(ATR_FILE, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                sym = str(row[0]).strip().upper()
                if not sym: continue
                # Sütun sırası: Sembol, Periyod, Birim, Anlık Fiyat, ATR 5dk, ATR 60dk, ATR 120dk, ATR 240dk, ATR Günlük, ATR Haftalık, ATR Ortalama
                vals = list(row)
                atr[sym] = {
                    'price':        _safe(vals[3]),
                    'atr_5dk':      _safe(vals[4]),
                    'atr_60dk':     _safe(vals[5]),
                    'atr_120dk':    _safe(vals[6]),
                    'atr_240dk':    _safe(vals[7]),
                    'atr_gunluk':   _safe(vals[8]),
                    'atr_haftalik': _safe(vals[9]),
                    'atr_ort':      _safe(vals[10]),
                }
            wb.close()
        except Exception as e:
            slog.warning(f'[ATR] okuma hatası: {e}')

    # Destek_Direc_Seviyeleri.xlsx
    if os.path.exists(DD_FILE):
        try:
            wb = load_workbook(DD_FILE, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                sym = str(row[0]).strip().upper()
                if not sym: continue
                # Sütun sırası: Sembol, Periyod, Birim, Hisse Fiyatı, Yakın Destek, Yakın Direnç, Orta Destek, Orta Direnç, Uzak Destek, Uzak Direnç, Durum, Trend Gücü
                vals = list(row)
                dd[sym] = {
                    'price2':    _safe(vals[3]),
                    'yakin_sup': _safe(vals[4]),
                    'yakin_res': _safe(vals[5]),
                    'orta_sup':  _safe(vals[6]),
                    'orta_res':  _safe(vals[7]),
                    'sup':       _safe(vals[8]),
                    'res':       _safe(vals[9]),
                    'durum':     str(vals[10] or '').strip(),
                    'trend':     str(vals[11] or '').strip(),
                }
            wb.close()
        except Exception as e:
            slog.warning(f'[DD] okuma hatası: {e}')

    # Birleştir
    syms = set(atr) | set(dd)
    result = {}
    for sym in syms:
        entry = {'ts': int(time.time())}
        entry.update(atr.get(sym, {}))
        entry.update(dd.get(sym, {}))
        # Fiyat: ATR dosyasındaki Anlik Fiyat öncelikli, yoksa DD dosyasındaki
        if entry.get('price') is None and entry.get('price2') is not None:
            entry['price'] = entry['price2']
        entry.pop('price2', None)
        result[sym] = entry

    # ── DDE Excel (MatriksIQ) anlık fiyat override ──────────────────────────
    # ATR_Sonuc.xlsx günde bir üretilir; DDE Excel ise canlı.
    # Piyasa saatlerinde tüm sembollerin price alanını anlık değerle güncelle.
    try:
        from price_reader import get_all_prices
        live = get_all_prices()
        for sym, lp in live.items():
            if lp and lp > 0:
                if sym in result:
                    result[sym]['price'] = round(lp, 4)
                    result[sym]['live'] = True
                else:
                    result[sym] = {'price': round(lp, 4), 'live': True,
                                   'ts': int(time.time())}
    except Exception as e:
        slog.debug(f'[DDE] anlık fiyat override atlandı: {e}')

    return result

def _safe(v):
    if v is None: return None
    try: return round(float(v), 4)
    except: return None


# Grid config — bir kez oku, cache'le
_GRID_CFG = None
def _load_grid_cfg():
    global _GRID_CFG
    if _GRID_CFG is None:
        try:
            with open(os.path.join(BASE_DIR, 'grid_analysis_config.json'),
                      'r', encoding='utf-8') as f:
                _GRID_CFG = json.load(f)
        except Exception:
            _GRID_CFG = {'capital': 1400000, 'safety_buffer': 0.9,
                         'commission_rate': 0.0001}
    return _GRID_CFG


def _bist_tick(price):
    """BİST fiyat adımı (web getBistTickSize ile birebir)."""
    if not price or price <= 0: return 0.01
    if price < 20:   return 0.01
    if price < 50:   return 0.02
    if price < 100:  return 0.05
    if price < 250:  return 0.10
    if price < 500:  return 0.25
    if price < 1000: return 0.50
    if price < 2500: return 1.00
    return 2.50


def _user_grid_capital():
    """Web hesaplayıcının (settings.gridCalc) sermayesi; yoksa config."""
    try:
        gc = firebase_get('settings/gridCalc')
        if isinstance(gc, dict) and gc.get('capital'):
            return float(gc['capital'])
    except Exception:
        pass
    return _load_grid_cfg().get('capital', 1400000)


def _recompute_grid_live(data, price):
    """
    Anlık fiyatla grid metriklerini WEB HESAPLAYICI (calcGridBot) algoritmasıyla
    yeniden hesaplar — N optimizasyonu + tick hizalama. Böylece web ve LCD birebir
    örtüşür. Sabit: support, resistance, atr (akşam analizi). Sermaye: gridCalc.
    """
    import math
    try:
        support    = float(data.get('support', 0) or 0)
        resistance = float(data.get('resistance', 0) or 0)
        atr        = float(data.get('atr', 0) or 0)
        if price <= 0 or atr <= 0 or resistance <= support:
            data['price'] = round(price, 4)
            return data

        cfg     = _load_grid_cfg()
        capital = _user_grid_capital()
        comm    = cfg.get('commission_rate', 0.0001)

        rng = resistance - support
        mid = (support + resistance) / 2.0
        tick = _bist_tick(price if price > 0 else mid)
        ref  = price if (support < price < resistance) else mid
        eff_cap = capital * 0.90
        min_spacing = max(tick, mid * 0.005)
        max_n = min(200, int(rng / min_spacing))
        min_n_atr = math.ceil(rng / (atr * 0.15)) if atr > 0 else 2
        narrow = min_n_atr > max_n
        min_n = max(2, max_n) if narrow else max(2, min_n_atr)
        if max_n < 2:
            data['price'] = round(price, 4)
            return data

        best = None
        for N in range(min_n, max_n + 1):
            d_raw = rng / N
            d = max(tick, math.ceil(d_raw / tick) * tick)
            sell_lv  = max(1, int((resistance - ref) / d))
            buy_lv   = max(1, int((ref - support) / d))
            avg_down = (ref + support) / 2.0
            cap_lot  = sell_lv * ref + buy_lv * avg_down
            if cap_lot <= 0:
                continue
            lots = int(eff_cap / cap_lot)
            if lots < 1:
                continue
            gross = lots * d
            comm_c = lots * comm * (mid + (mid + d))
            net = gross - comm_c
            if net <= 0:
                continue
            daily = (atr / d) * net
            used  = lots * cap_lot
            if best is None or daily > best['daily']:
                best = {'N': N, 'd': d, 'lots': lots, 'daily': daily,
                        'used': used, 'sell': sell_lv, 'buy': buy_lv}

        data['price'] = round(price, 4)
        if best:
            # Al/Sat dağılımı: web renderGridResults EKRAN mantığıyla birebir.
            # N+1 seviye, dExact=range/N (yuvarlanmamış), her seviye tick'e
            # yuvarlanıp ref'e göre al/sat/mevcut sınıflandırılır.
            N = best['N']
            d_exact = rng / N
            buy_cnt = 0
            sell_cnt = 0
            for i in range(N + 1):
                p_lvl = round((support + i * d_exact) / tick) * tick
                if abs(p_lvl - ref) < tick / 2:
                    continue  # mevcut fiyat seviyesi ('at')
                elif p_lvl > ref:
                    sell_cnt += 1
                else:
                    buy_cnt += 1

            data['grid_interval'] = round(best['d'], 2)
            data['lots']          = best['lots']
            data['sell_grids']    = sell_cnt
            data['buy_grids']     = buy_cnt
            data['total_grids']   = buy_cnt + sell_cnt
            data['daily_profit']  = round(best['daily'], 0)
            data['capital_used']  = round(best['used'], 0)
            data['pct_up']        = round((resistance - ref) / ref * 100, 2)
            data['pct_down']      = round((ref - support)    / ref * 100, 2)
            data['live_recalc']   = True
    except Exception:
        data['price'] = round(price, 4)
    return data

# ---------------------------------------------------------------------------
# Firebase
# ---------------------------------------------------------------------------

def firebase_put(path, data):
    try:
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(
            f'{FIREBASE_URL}/{path}.json', data=body, method='PUT',
            headers={'Content-Type': 'application/json'})
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        slog.warning(f'[Firebase] hata: {e}')

def firebase_get(path):
    try:
        req = urllib.request.Request(
            f'{FIREBASE_URL}/{path}.json',
            headers={'Content-Type': 'application/json'})
        resp = urllib.request.urlopen(req, timeout=15)
        return json.loads(resp.read().decode())
    except Exception as e:
        slog.warning(f'[Firebase] GET hata: {e}')
        return None

def get_server_ip():
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None):
            ip = info[4][0]
            if ip.startswith('100.'):
                return ip
    except: pass
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]; s.close(); return ip
    except: return '127.0.0.1'

# ---------------------------------------------------------------------------
# Arka plan thread: Excel oku → belleği güncelle → Firebase'e push
# ---------------------------------------------------------------------------

def _bg_loop():
    while True:
        try:
            stocks = read_excel()
            ts = int(time.time())
            with _lock:
                _stocks.clear()
                _stocks.update(stocks)
                global _stocks_ts
                _stocks_ts = ts
            # Firebase'e yaz
            firebase_put('gridtracker/stocks', stocks)
            firebase_put('gridtracker/stocks_ts', ts)
        except Exception as e:
            slog.warning(f'[BG] hata: {e}')
        time.sleep(60)

def _push_queue_loop():
    """Her 10 saniyede bir Firebase pushQueue kontrol et."""
    while True:
        try:
            _check_push_queue()
        except Exception as e:
            slog.warning(f'[PushQueue] hata: {e}')
        time.sleep(10)

# ---------------------------------------------------------------------------
# Push Bildirimleri
# ---------------------------------------------------------------------------

def _load_vapid():
    try:
        with open(VAPID_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        slog.warning(f'[Push] VAPID yüklenemedi: {e}')
        return None

def send_push_to_all(title, body, tag='gridtracker'):
    keys = _load_vapid()
    if not keys:
        slog.warning('[Push] VAPID anahtarı yok, bildirim atlandı.')
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        slog.warning('[Push] pywebpush yüklü değil.')
        return
    try:
        req = urllib.request.Request(
            f'{FIREBASE_URL}/gridtracker/pushSubscriptions.json')
        resp = urllib.request.urlopen(req, timeout=10)
        subs = json.loads(resp.read().decode())
    except Exception as e:
        slog.warning(f'[Push] Aboneler alınamadı: {e}')
        return
    if not subs or not isinstance(subs, dict):
        slog.warning('[Push] Kayıtlı abone yok — telefon uygulamayı kapatmış olabilir, yeniden abone ol.')
        return
    payload = json.dumps({'title': title, 'body': body, 'tag': tag})
    sent = 0
    for sub_key, sub_data in subs.items():
        try:
            if isinstance(sub_data, str):
                sub_data = json.loads(sub_data)
            webpush(
                subscription_info=sub_data,
                data=payload,
                vapid_private_key=keys['privateKey'],
                vapid_claims=VAPID_CLAIMS,
                ttl=86400,
                headers={'urgency': 'high'}
            )
            sent += 1
            slog.info(f'[Push] Gönderildi: {sub_key[:12]}…')
        except WebPushException as e:
            status = getattr(e.response, 'status_code', None) if e.response else None
            if status == 410:
                # Subscription kalıcı olarak silindi (410) — Firebase'den kaldır
                try:
                    del_req = urllib.request.Request(
                        f'{FIREBASE_URL}/gridtracker/pushSubscriptions/{sub_key}.json',
                        method='DELETE')
                    urllib.request.urlopen(del_req, timeout=5)
                    slog.warning(f'[Push] Kalıcı silinen abone kaldırıldı (HTTP 410): {sub_key[:12]}…')
                except Exception:
                    pass
            else:
                slog.warning(f'[Push] WebPush hatası [{sub_key[:12]}…]: {e}')
        except Exception as e:
            slog.warning(f'[Push] Beklenmeyen hata [{sub_key[:12]}…]: {e}')
    slog.info(f'[Push] Tamamlandı: {sent}/{len(subs)} abone başarılı — {title}')

def _check_push_queue():
    """Firebase pushQueue'daki bildirimleri gönder ve sil."""
    try:
        req = urllib.request.Request(
            f'{FIREBASE_URL}/gridtracker/pushQueue.json')
        resp = urllib.request.urlopen(req, timeout=10)
        queue = json.loads(resp.read().decode())
        if not queue or not isinstance(queue, dict):
            return
        for tag, item in queue.items():
            if not isinstance(item, dict):
                continue
            title = item.get('title', 'GridTracker')
            body  = item.get('body', '')
            send_push_to_all(title, body, tag)
            # Sil
            del_req = urllib.request.Request(
                f'{FIREBASE_URL}/gridtracker/pushQueue/{tag}.json',
                method='DELETE')
            urllib.request.urlopen(del_req, timeout=5)
    except Exception as e:
        slog.warning(f'[PushQueue] hata: {e}')

# ---------------------------------------------------------------------------
# ntfy.sh Push
# ---------------------------------------------------------------------------

def send_ntfy(title, body, priority=3, tags=None, alert=False):
    """ntfy.sh üzerinden push bildirimi gönderir (JSON body — Türkçe + emoji destekli).
    alert=True → kritik (ÇIK) topic (telefonda alarm sesi). Başlığa [ACİL] eklenir."""
    import http.client, json
    try:
        if alert and '[ACİL]' not in title:
            title = f'[ACİL] {title}'
        payload = {
            'topic':    NTFY_TOPIC_ALERT if alert else NTFY_TOPIC,
            'title':    title,
            'message':  body,
            'priority': priority,   # 1=min 2=low 3=default 4=high 5=max
        }
        if tags:
            payload['tags'] = tags if isinstance(tags, list) else [tags]
        data = json.dumps(payload, ensure_ascii=False).encode('utf-8')
        conn = http.client.HTTPSConnection('ntfy.sh', timeout=10)
        conn.request('POST', '/', body=data,
                     headers={'Content-Type': 'application/json'})
        resp = conn.getresponse()
        conn.close()
        slog.info(f'[ntfy] Gönderildi: {title} (HTTP {resp.status})')
    except Exception as e:
        slog.warning(f'[ntfy] Hata: {e}')


# ---------------------------------------------------------------------------
# Pozisyon Verdict Hesaplama (paylaşılan fonksiyon)
# ---------------------------------------------------------------------------

def _grid_verdict(gs, trend_dir, price, g_sup, g_res):
    """Grid pozisyon kararı (ortak mantık).
    Kullanıcı kuralı: GRID BİR ROTASYON STRATEJİSİ DEĞİL — taahhüt + zaman ister.
    Bu yüzden 'başkası daha iyi → değiştir' dürtüsü KALDIRILDI. Değiştir/çık
    sinyali YALNIZCA: (a) fiyat grid aralığını kırdı, (b) grid skoru çöktü.
    Döner: (verdict, title, reason)"""
    if gs <= 0:
        return 'dikkat', 'VERI YOK', 'Analiz bekleniyor...'
    # 1) Fiyat grid aralığının ALTINDA → grid kırıldı (düşüşte sıkıştı, TUPRS gibi)
    if g_sup > 0 and 0 < price < g_sup:
        return ('cik', 'ÇIK',
                f'Fiyat grid aralığının ALTINA indi ({price:.2f} < destek '
                f'{g_sup:.2f}). Grid kırıldı — çık, yeni grid kur.')
    # 2) Fiyat grid aralığının ÜSTÜNDE → grid tamamlandı (yukarı kaçtı = kâr)
    if g_res > 0 and price > g_res:
        return ('dikkat', 'DEĞİŞTİR',
                f'Fiyat grid aralığının ÜSTÜNE çıktı ({price:.2f} > direnç '
                f'{g_res:.2f}). Grid tamamlandı — kârı al, yeni grid kur.')
    # 3) Grid skoru çöktü → hisse salınmayı bıraktı / trende girdi
    if gs < 3.5:
        return 'cik', 'ÇIK', f'Grid skoru çok düşük ({gs:.1f}/10). Hisse grid için verimsiz.'
    # 4) Zayıflama → İZLE (switch DEĞİL; grid'e zaman ver, aralığı bekle)
    if gs < 4 or (trend_dir == 'falling' and gs < 5):
        return ('dikkat', 'DİKKAT',
                f'Grid skoru zayıfladı ({gs:.1f}/10). İzle — aralık kırılırsa çık.')
    # 5) Sağlıklı
    t = ' Skor yükseliyor.' if trend_dir == 'rising' else ''
    return 'devam', 'DEVAM ET', f'Skor iyi ({gs:.1f}/10), fiyat grid aralığında.{t}'


def compute_pos_verdict():
    """
    Firebase verilerinden aktif açık pozisyonun karar skorunu hesaplar.
    Döner: dict (symbol, gridScore, finalScore, verdict, title, reason, trend)
    ya da None (veri yoksa / hesaplanamıyorsa).
    """
    try:
        tp = firebase_get('gridtracker/todayProfit')
        if not tp:
            return None
        open_pos_all = tp.get('openPositions', {})
        if not open_pos_all:
            return None

        # botSymbols — Firebase settings'ten al
        settings_fb = firebase_get('gridtracker/settings') or {}
        bot_syms    = [s.upper() for s in (settings_fb.get('botSymbols') or [])]
        open_pos    = {k: v for k, v in open_pos_all.items() if k in bot_syms} if bot_syms else open_pos_all
        if not open_pos:
            return None

        # En fazla lotlu sembol
        stocks_raw = firebase_get('gridtracker/stocks') or {}
        best_sym, best_qty = '', 0
        for sym, pl in open_pos.items():
            q = sum(p.get('execQty', 0) for p in pl)
            if q > best_qty:
                best_qty, best_sym = q, sym
        if not best_sym:
            return None

        pv_sym = best_sym

        # gridRecActive (aktif sembol) → yoksa gridRec (en iyi öneri)
        gr_active = firebase_get('gridtracker/gridRecActive') or {}
        rec_sym   = (gr_active.get('symbol') or '').replace('.IS', '').upper()
        if rec_sym != pv_sym:
            gr_active = firebase_get('gridtracker/gridRec') or {}
            rec_sym   = (gr_active.get('symbol') or '').replace('.IS', '').upper()
        gs = float(gr_active.get('grid_score',  0) or 0) if rec_sym == pv_sym else 0.0
        fs = float(gr_active.get('final_score', 0) or 0) if rec_sym == pv_sym else 0.0

        # Trend — son 3 günlük skor geçmişi
        trend_dir = 'stable'
        hist = firebase_get(f'gridtracker/scoreHistory/{pv_sym}') or []
        if isinstance(hist, list) and len(hist) >= 3:
            last3 = [float(h.get('gs', 0) or 0) for h in hist[-3:]]
            delta = last3[2] - last3[0]
            if   delta >  0.5: trend_dir = 'rising'
            elif delta < -0.5: trend_dir = 'falling'

        # GRID ARALIĞI: kullanıcının kurduğu grid (gridCalc) öncelikli; yoksa
        # sistemin hesapladığı destek/direnç (gridRecActive).
        price = float(stocks_raw.get(pv_sym, {}).get('price', 0) or 0)
        gc = settings_fb.get('gridCalc') or {}
        if (gc.get('symbol', '') or '').upper() == pv_sym and gc.get('support') and gc.get('resistance'):
            g_sup, g_res = float(gc.get('support') or 0), float(gc.get('resistance') or 0)
        else:
            g_sup = float(gr_active.get('support', 0) or 0)
            g_res = float(gr_active.get('resistance', 0) or 0)

        verdict, title, reason = _grid_verdict(gs, trend_dir, price, g_sup, g_res)

        return {'symbol': pv_sym, 'gridScore': round(gs, 1), 'finalScore': round(fs, 3),
                'verdict': verdict, 'title': title, 'reason': reason, 'trend': trend_dir}
    except Exception as e:
        slog.warning(f'[posVerdict] hata: {e}')
        return None


# ---------------------------------------------------------------------------
# Verdict Değişiklik Monitörü
# ---------------------------------------------------------------------------

def _verdict_monitor_loop():
    """
    Her 90 saniyede posVerdict'i hesaplar.
    verdict veya sembol değişince push bildirimi gönderir.
    Son durum Firebase gridtracker/verdictState altında saklanır.
    """
    time.sleep(45)   # Server tam ayağa kalksın
    while True:
        try:
            pv = compute_pos_verdict()
            # Veri yoksa veya anlamlı bir durum yoksa atla
            if not pv or (pv['verdict'] == 'dikkat' and pv['title'] == 'VERI YOK'):
                time.sleep(90)
                continue

            sym        = pv['symbol']
            new_verdict = pv['verdict']
            new_title   = pv['title']

            # Firebase'deki son kayıtlı durum
            last       = firebase_get('gridtracker/verdictState') or {}
            last_sym   = last.get('symbol', '')
            last_title = last.get('title', '')   # title bazlı karşılaştırma (DEĞİŞTİR? ayrı sayılır)

            if last_sym != sym or last_title != new_title:
                emoji_map = {'devam': '✅', 'dikkat': '⚠️', 'cik': '🔴'}
                prio_map  = {'devam': 3,   'dikkat': 4,     'cik': 5}
                em = emoji_map.get(new_verdict, '•')
                push_title = f'{em} {sym} → {new_title} {em}'
                push_body  = pv['reason']

                slog.info(f'[Verdict] Değişiklik: {last_sym}/{last_title} → {sym}/{new_title}')
                threading.Thread(
                    target=send_ntfy,
                    args=(push_title, push_body,
                          prio_map.get(new_verdict, 3)),
                    kwargs={'alert': new_verdict == 'cik'},   # ÇIK → ACİL topic
                    daemon=True
                ).start()

                # Yeni durumu kaydet
                firebase_put('gridtracker/verdictState', {
                    'symbol':  sym,
                    'verdict': new_verdict,
                    'title':   pv['title'],
                    'reason':  pv['reason'],
                    'ts':      int(time.time()),
                })
        except Exception as e:
            slog.warning(f'[VerdictMonitor] hata: {e}')
        time.sleep(90)


def _market_open_now():
    """BIST açık mı: hafta içi 10:00–18:15 (yaklaşık)."""
    t = time.localtime()
    if t.tm_wday >= 5:            # Cmt/Pazar
        return False
    mins = t.tm_hour * 60 + t.tm_min
    return 10 * 60 <= mins <= 18 * 60 + 15


def _grid_scan_loop():
    """Gün içi grid aday taraması — 'En İyi Grid Adayı' kartı dünkü veride
    kalmasın. Piyasa açıkken her 30 dk grid_analysis_auto.py --force çalıştırır
    (50 hisseyi yeniden tarar, sonucu json + Firebase'e yazar)."""
    time.sleep(60)   # server otursun
    SCAN_MIN = 30
    last_scan = 0.0
    script = os.path.join(BASE_DIR, 'grid_analysis_auto.py')
    while True:
        try:
            if _market_open_now() and (time.time() - last_scan >= SCAN_MIN * 60):
                if os.path.exists(script):
                    slog.info('[GridScan] Gün içi grid taraması başlıyor...')
                    r = subprocess.run([sys.executable, script, '--force'],
                                       cwd=BASE_DIR, capture_output=True, text=True,
                                       encoding='utf-8', errors='replace', timeout=300)
                    if r.returncode == 0:
                        slog.info('[GridScan] tamamlandı ✓')
                    else:
                        slog.warning(f'[GridScan] hata: {r.stderr[:200]}')
                last_scan = time.time()
        except Exception as e:
            slog.warning(f'[GridScan] hata: {e}')
        time.sleep(120)


# ---------------------------------------------------------------------------
# HTTP Handler
# ---------------------------------------------------------------------------

class Handler(SimpleHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def end_headers(self):
        # sw.js, manifest.json ve .html asla cache'lenmesin (taze sayfa garantisi)
        path = self.path.split('?')[0]
        if path in ('/sw.js', '/manifest.json') or path.endswith('.html') or path in ('/', ''):
            self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate')
            self.send_header('Pragma', 'no-cache')
        super().end_headers()

    def send_json(self, code, obj):
        body = json.dumps(obj, ensure_ascii=False).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        path = self.path.split('?')[0]
        # POST /api/notify
        if path == '/api/notify':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = self.rfile.read(length)
                data   = json.loads(body.decode('utf-8')) if body else {}
            except Exception:
                data = {}
            title = data.get('title', 'GridTracker')
            body  = data.get('body', '')
            tag   = data.get('tag', 'gridtracker')
            threading.Thread(
                target=send_push_to_all,
                args=(title, body, tag),
                daemon=True
            ).start()
            self.send_json(200, {'ok': True})
            return
        # POST /api/advisor/capital — Sermaye güncelle
        if path == '/api/advisor/capital':
            try:
                length = int(self.headers.get('Content-Length', 0))
                body   = self.rfile.read(length)
                data   = json.loads(body.decode('utf-8')) if body else {}
                capital = float(data.get('capital', 0))
                if capital < 0:
                    self.send_json(400, {'error': 'Geçersiz sermaye'}); return
                state_path = os.path.join(BASE_DIR, 'Gunluk_Al_Sat', 'state.json')
                if os.path.exists(state_path):
                    with open(state_path, 'r', encoding='utf-8') as f:
                        st = json.load(f)
                    st['capital'] = capital
                    with open(state_path, 'w', encoding='utf-8') as f:
                        json.dump(st, f, ensure_ascii=False, indent=2)
                    # ── Anında lot yeniden hesapla (tam analiz gerekmez) ──
                    # Mevcut top_picks fiyatlarıyla yeni sermayeye göre lot güncelle
                    try:
                        result_path = os.path.join(BASE_DIR, 'advisor_result.json')
                        if os.path.exists(result_path):
                            with open(result_path, 'r', encoding='utf-8') as f:
                                res = json.load(f)
                            res['capital'] = capital
                            comm = st.get('commission_rate', 0.0001)
                            eff  = capital * 0.90
                            for p in res.get('top_picks', []):
                                px   = p.get('price', 0)
                                if px <= 0: continue
                                lots      = int(eff / px) if px else 0
                                lots_main = int(eff * 0.60 / px) if px else 0
                                lots_dip  = int(eff * 0.25 / px) if px else 0
                                sym = p.get('symbol','')
                                if sym:
                                    res.setdefault('lot_info', {})[sym] = {
                                        'lots': lots, 'lots_main': lots_main,
                                        'lots_dip': lots_dip,
                                        'dip_price': round(px * 0.97, 4),
                                        'price': px,
                                        'cost':       round(lots_main * px * (1+comm), 2),
                                        'cost_total': round(lots * px * (1+comm), 2),
                                    }
                            with open(result_path, 'w', encoding='utf-8') as f:
                                json.dump(res, f, ensure_ascii=False)
                            # Firebase lot_info da güncelle
                            try:
                                import requests as _rq
                                _fb = 'https://grid-tracker-73ed2-default-rtdb.europe-west1.firebasedatabase.app/gridtracker/advisor'
                                _rq.patch(_fb+'.json', json={'capital': capital, 'lot_info': res.get('lot_info',{})}, timeout=6)
                            except Exception:
                                pass
                    except Exception as _e:
                        slog.debug(f'lot recalc: {_e}')
                    self.send_json(200, {'ok': True, 'capital': capital})
                else:
                    self.send_json(404, {'error': 'state.json bulunamadı'})
            except Exception as e:
                self.send_json(500, {'error': str(e)})
            return

        self.send_json(405, {'error': 'Method Not Allowed'})

    def do_GET(self):
        if self.path in ('/', ''):
            self.send_response(302)
            self.send_header('Location', '/bist_tracker.html')
            self.end_headers()
            return

        path  = self.path.split('?')[0]
        parts = path.strip('/').split('/')

        # /api/stock/{SYM}
        if len(parts) == 3 and parts[0] == 'api' and parts[1] == 'stock':
            sym = parts[2].upper()
            with _lock:
                data = dict(_stocks.get(sym) or {})
            # Anlık fiyatı her durumda price_reader'dan çek (DDE/cache/Yahoo).
            # Sembol ATR_Sonuc.xlsx'te (bot sembolleri) olmasa bile (ör. ALARK
            # gibi BIST50 grid adayları) en azından fiyat döndürülür.
            try:
                from price_reader import get_price
                lp, src = get_price(sym)
                if lp and lp > 0:
                    data['price'] = round(lp, 4)
                    data['live']  = (src == 'excel')
                    data['ts']    = int(time.time())
            except Exception:
                pass
            if data.get('price'):
                self.send_json(200, data)
            else:
                self.send_json(404, {'error': f'{sym} fiyatı alınamadı'})
            return

        # /api/all
        if path == '/api/all':
            with _lock:
                self.send_json(200, {'stocks': dict(_stocks), 'ts': _stocks_ts})
            return

        # /api/health
        if path == '/api/health':
            with _lock:
                self.send_json(200, {'ok': True, 'count': len(_stocks), 'ts': _stocks_ts})
            return

        # /api/grid-data — LCD için realNet, openPosCount, todayProfit
        if path == '/api/grid-data':
            bt_file = os.path.join(BASE_DIR, 'bist_tracker.html')
            realNet = 0.0
            openPosCount = 0
            todayProfit = 0.0
            todayOverall = 0.0
            if os.path.exists(bt_file):
                try:
                    with open(bt_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                    import re
                    m = re.search(r'"realNet":\s*([-\d.]+)', content)
                    if m: realNet = float(m.group(1))
                    m = re.search(r'"openPosCount":\s*(\d+)', content)
                    if m: openPosCount = int(m.group(1))
                    m = re.search(r'"todayOverall":\s*([-\d.]+)', content)
                    if m: todayOverall = float(m.group(1))
                    m = re.search(r'"netProfit":\s*([-\d.]+)', content)
                    if m: todayProfit = float(m.group(1))
                except Exception as e:
                    slog.warning(f'[GridData] okuma hatası: {e}')
            self.send_json(200, {
                'realNet': realNet,
                'openPosCount': openPosCount,
                'todayProfit': todayProfit,
                'todayOverall': todayOverall
            })
            return

        # /api/all-data — LCD için TÜM VERİ (tek istek, Firebase'e gerek yok)
        if path == '/api/all-data':
            bt_file = os.path.join(BASE_DIR, 'bist_tracker.html')
            data = {
                'today': '', 'todayOverall': 0.0,
                'realNet': 0.0, 'openPosCount': 0,
                'todayNet': 0.0, 'todayGross': 0.0, 'todayComm': 0.0,
                'todayBuys': 0, 'todaySells': 0,
                'totalProfit': 0.0, 'portfolioDiff': 0.0,
                'currentMonthNet': 0.0,
                'botSymbols': [], 'monthlyKar': [], 'openPositions': [],
                'dailyLog': {},
                'posMonitor': {},
                'posVerdict': {}
            }
            if os.path.exists(bt_file):
                try:
                    with open(bt_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                    import re

                    # window.__GRID_DATA__ bloğunu çıkar — non-greedy + DOTALL
                    m = re.search(r'window\.__GRID_DATA__\s*=\s*(\{.*?})\s*;', content, re.DOTALL)
                    if not m:
                        m = re.search(r'window\.__GRID_DATA__\s*=\s*(\{.*)', content)
                    if m:
                        try:
                            json_str = m.group(1)
                            # JavaScript temizleme
                            json_str = re.sub(r'^\s*//.*', '', json_str, flags=re.MULTILINE)
                            json_str = re.sub(r',(\s*[}\]])', r'\1', json_str)
                            try:
                                gd = json.loads(json_str)
                            except json.JSONDecodeError as e:
                                # Kırpılmış olabilir — sonunda fazlalık var, tamamlanana kadar kırp
                                depth = 0
                                last_ok = 0
                                for i, c in enumerate(json_str):
                                    if c == '{': depth += 1
                                    elif c == '}':
                                        depth -= 1
                                        if depth == 0:
                                            last_ok = i + 1
                                            break
                                if last_ok > 0:
                                    gd = json.loads(json_str[:last_ok])
                                else:
                                    gd = {}

                            data['today'] = gd.get('today', '')
                            data['todayOverall'] = gd.get('todayOverall', 0.0)

                            # realNet & openPosCount — FIREBASE'DEN ANLIK AL (canlı veri)
                            tp = firebase_get('gridtracker/todayProfit')
                            if tp:
                                total_net = tp.get('totalNet', 0.0)
                                open_pos_all = tp.get('openPositions', {})

                                bot_syms = data.get('botSymbols', [])
                                if not bot_syms:
                                    settings = gd.get('settings', {})
                                    bot_syms = settings.get('botSymbols', [])
                                open_pos = {k: v for k, v in open_pos_all.items() if k in bot_syms}

                                unrealized = 0.0
                                stocks_raw = firebase_get('gridtracker/stocks')
                                if stocks_raw:
                                    for sym, pos_list in open_pos.items():
                                        cur_price = stocks_raw.get(sym, {}).get('price', 0)
                                        if cur_price <= 0: continue
                                        for pos in pos_list:
                                            qty = pos.get('execQty', 0)
                                            px = pos.get('execPrice', 0)
                                            exec_amt = pos.get('execAmount', 0)
                                            comm = pos.get('commission', 0)
                                            comm_rate = exec_amt > 0 and comm / exec_amt or 0.0001
                                            sell_comm = cur_price * qty * comm_rate
                                            unrealized += (cur_price - px) * qty - sell_comm
                                data['realNet'] = total_net + unrealized
                                open_count = sum(len(v) for v in open_pos.values())
                                data['openPosCount'] = open_count

                                # posMonitor — en fazla lotlu botSymbol (LCD 5. sayfa)
                                pos_mon = {'symbol': '', 'qty': 0, 'avgCost': 0.0,
                                           'curPrice': 0.0, 'unrealPnl': 0.0, 'posCount': 0}
                                best_sym, best_qty = '', 0
                                for sym, pl in open_pos.items():
                                    q = sum(p.get('execQty', 0) for p in pl)
                                    if q > best_qty:
                                        best_qty, best_sym = q, sym
                                if best_sym and best_qty > 0:
                                    pl = open_pos[best_sym]
                                    avg_c = sum(p.get('execQty', 0) * p.get('execPrice', 0) for p in pl) / best_qty
                                    c_px = stocks_raw.get(best_sym, {}).get('price', 0) if stocks_raw else 0
                                    cr = 0.0001
                                    if pl and pl[0].get('execAmount', 0) > 0:
                                        cr = pl[0].get('commission', 0) / pl[0].get('execAmount', 1)
                                    u_pnl = (c_px - avg_c) * best_qty - c_px * best_qty * cr if c_px > 0 else 0.0
                                    pos_mon = {'symbol': best_sym, 'qty': best_qty,
                                               'avgCost': round(avg_c, 2), 'curPrice': round(c_px, 2),
                                               'unrealPnl': round(u_pnl, 0), 'posCount': len(pl)}
                                data['posMonitor'] = pos_mon

                                # posVerdict — ortak compute_pos_verdict ile (grid
                                # aralığı kuralı + 'başkası daha iyi' dürtüsü yok).
                                pv_sym = pos_mon.get('symbol', '')
                                pv = compute_pos_verdict() or {
                                    'symbol': pv_sym, 'gridScore': 0.0, 'finalScore': 0.0,
                                    'verdict': 'dikkat', 'title': 'VERI YOK',
                                    'reason': 'Analiz bekleniyor...', 'trend': 'stable'}
                                data['posVerdict'] = pv
                            else:
                                data['realNet'] = gd.get('realNet', 0.0)
                                data['openPosCount'] = gd.get('openPosCount', 0)

                            # dailyLog (bugünün net/gross/comm/buys/sells)
                            dl = gd.get('dailyLog', {})
                            today_key = gd.get('today', '')
                            today_dl = dl.get(today_key, {}) if today_key else {}
                            data['todayNet'] = today_dl.get('netProfit', 0.0)
                            data['todayGross'] = today_dl.get('grossProfit', 0.0)
                            data['todayComm'] = today_dl.get('commission', 0.0)
                            data['todayBuys'] = today_dl.get('buys', 0)
                            data['todaySells'] = today_dl.get('sells', 0)

                            # totalProfit & portfolioDiff
                            data['totalProfit'] = gd.get('totalProfit', 0.0)
                            data['portfolioDiff'] = gd.get('portfolioDiff', 0.0)

                            # portfolioDiff — bugünün overall farkı
                            oh = gd.get('overallHistory', [])
                            today_ov = gd.get('todayOverall', 0.0)
                            today_cap_sum = 0.0
                            birikim = gd.get('birikimTx', [])
                            for tx in birikim:
                                d = tx.get('date', '')
                                if d == today_key and not tx.get('exclude', False) and not tx.get('trackPayment', False):
                                    today_cap_sum += tx.get('amount', 0.0)
                            prev_oh = None
                            for h in sorted(oh, key=lambda x: x['date'], reverse=True):
                                if h['date'] < today_key:
                                    prev_oh = h
                                    break
                            if prev_oh:
                                perf_diff = (today_ov - prev_oh.get('amount', 0.0)) - today_cap_sum
                                data['portfolioDiff'] = perf_diff

                            # currentMonthNet — aylık kar (bu ay + geçmiş aylık fark)
                            import datetime
                            now = datetime.datetime.now()
                            cur_year = now.year
                            cur_month = now.month
                            cur_month_key = f'{cur_year}-{cur_month:02d}'
                            month_net = 0.0
                            for dk, dv in dl.items():
                                if dk.startswith(cur_month_key):
                                    month_net += dv.get('netProfit', 0.0)
                            data['currentMonthNet'] = month_net

                            # botSymbols — settings altında
                            settings = gd.get('settings', {})
                            data['botSymbols'] = settings.get('botSymbols', [])

                            # monthlyKar
                            data['monthlyKar'] = gd.get('monthlyKar', [])

                            # monthlyTradesNet — son 3 ayın işlem netleri (dailyLog'dan)
                            import datetime
                            now2 = datetime.datetime.now()
                            cur_year2 = now2.year
                            cur_month2 = now2.month
                            turkish_months = {1:'OCAK',2:'SUBAT',3:'MART',4:'NISAN',5:'MAYIS',6:'HAZIRAN',
                                              7:'TEMMUZ',8:'AGUSTOS',9:'EYLUL',10:'EKIM',11:'KASIM',12:'ARALIK'}
                            # Ayları topla
                            mons_sum = {}
                            for dk in dl.keys():
                                if isinstance(dk, str) and len(dk) >= 7:
                                    ym = dk[:7]
                                    if ym not in mons_sum:
                                        mons_sum[ym] = 0.0
                                    mons_sum[ym] += dl[dk].get('netProfit', 0.0)
                            # Son 3 ayı sırala
                            sorted_mons = sorted(mons_sum.items(), key=lambda x: x[0], reverse=True)
                            monthly_trades_net = []
                            for ym, profit in sorted_mons[:3]:
                                parts = ym.split('-')
                                m = int(parts[1])
                                trname = turkish_months.get(m, ym)
                                monthly_trades_net.append({'key': ym, 'trname': trname, 'profit': int(profit)})
                            data['monthlyTradesNet'] = monthly_trades_net

                            # openPositions
                            data['openPositions'] = gd.get('openPositions', {})

                            # dailyLog (sadece aylık özet — ESP32 tampon sınırı için)
                            data['dailyLog'] = {}

                            slog.info(f'[AllData] today={data["today"]}')
                        except json.JSONDecodeError as e:
                            slog.warning(f'[AllData] JSON parse hatası: {e}')
                except Exception as e:
                    slog.warning(f'[AllData] okuma hatası: {e}')
            self.send_json(200, data)
            return

        # /api/grid-analysis — Grid bot analiz sonucu (ESP32 LCD 3. sayfa)
        if path == '/api/grid-analysis':
            result_file = os.path.join(BASE_DIR, 'grid_analysis_result.json')
            if os.path.exists(result_file):
                try:
                    with open(result_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    # Anlık DDE fiyatıyla TÜM grid hesabını yeniden hesapla
                    # (lot, grid sayısı, kar, emir% anlık fiyata göre güncellenir;
                    #  destek/direnç/ATR/grid aralığı akşam analizinden sabit kalır)
                    try:
                        sym = (data.get('symbol') or '').replace('.IS', '').upper()
                        if sym:
                            from price_reader import get_price
                            lp, src = get_price(sym)
                            if lp and lp > 0:
                                data = _recompute_grid_live(data, lp)
                                data['price_live'] = (src == 'excel')
                    except Exception as e:
                        slog.debug(f'[GridAnaliz] anlık hesap atlandı: {e}')
                    self.send_json(200, data)
                except Exception as e:
                    slog.warning(f'[GridAnaliz] okuma hatası: {e}')
                    self.send_json(500, {'error': str(e)})
            else:
                self.send_json(404, {'error': 'Henuz analiz yapilmadi'})
            return

        # /api/advisor — Günlük Sermaye Danışmanı sonucu
        if path == '/api/advisor' or path == '/advisor_result.json':
            result_file = os.path.join(BASE_DIR, 'advisor_result.json')
            if os.path.exists(result_file):
                try:
                    with open(result_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    self.send_json(200, data)
                except Exception as e:
                    self.send_json(500, {'error': str(e)})
            else:
                self.send_json(404, {'error': 'Henuz analiz yapilmadi'})
            return

        # Statik dosyalar
        super().do_GET()

    def translate_path(self, path):
        import posixpath, urllib.parse
        path = posixpath.normpath(urllib.parse.unquote(path.split('?')[0]))
        result = BASE_DIR
        for part in path.split('/'):
            if part in ('', os.curdir, os.pardir): continue
            result = os.path.join(result, part)
        return result

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    os.chdir(BASE_DIR)
    # İlk okuma
    stocks = read_excel()
    with _lock:
        _stocks.update(stocks)
        _stocks_ts = int(time.time())
    slog.info(f'[GridTracker] {len(_stocks)} hisse yüklendi.')
    # Firebase'e yaz (IP + ilk veri)
    ip = get_server_ip()
    threading.Thread(target=firebase_put, daemon=True,
        args=('gridtracker/serverInfo', {'ip': ip, 'port': PORT})).start()
    threading.Thread(target=firebase_put, daemon=True,
        args=('gridtracker/stocks', dict(_stocks))).start()
    # Arka plan döngüsü
    threading.Thread(target=_bg_loop,              daemon=True).start()
    # Verdict değişiklik monitörü — push bildirimi gönderir
    threading.Thread(target=_verdict_monitor_loop, daemon=True).start()
    # Gün içi grid aday taraması (30 dk, piyasa açıkken) — kart güncel kalsın
    threading.Thread(target=_grid_scan_loop, daemon=True).start()
    # Push queue → automation_server.pyw (port 5051) tarafından işleniyor
    # ThreadingHTTPServer: her istek ayrı thread'de işlenir. Bir istek (örn
    # Excel COM) takılsa bile diğer istekler (health, all-data) çalışmaya devam
    # eder → server tamamen donmaz, LCD/HTML veri almaya devam eder.
    server = ThreadingHTTPServer(('0.0.0.0', PORT), Handler)
    server.daemon_threads = True
    slog.info(f'[GridTracker] http://localhost:{PORT}/bist_tracker.html')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass

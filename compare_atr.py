import openpyxl, yfinance as yf, numpy as np, warnings
warnings.filterwarnings('ignore')

# Excel oku
wb = openpyxl.load_workbook('ATR_Sonuc.xlsx', data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
excel = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    sym = str(row[0]).strip().upper()
    excel[sym] = {'atr_gunluk': row[8], 'atr_ort': row[10], 'price': row[3]}
wb.close()

def calc_atr14(sym):
    try:
        df = yf.download(sym+'.IS', period='60d', interval='1d', progress=False, auto_adjust=True)
        if len(df) < 15: return None
        h = df['High'].squeeze()
        l = df['Low'].squeeze()
        c = df['Close'].squeeze()
        prev_c = c.shift(1)
        tr = np.maximum(h-l, np.maximum(abs(h-prev_c), abs(l-prev_c)))
        return float(tr.rolling(14).mean().iloc[-1])
    except:
        return None

header = f"{'Sembol':8s}  {'Fiyat':>8s}  {'Excel Gunluk':>12s}  {'Excel Ort':>9s}  {'YF ATR14':>8s}  {'Fark':>8s}  {'Fark%':>6s}"
print(header)
print('-' * 75)

diffs = []
for sym in sorted(excel.keys()):
    yf_atr = calc_atr14(sym)
    e = excel[sym]
    eg = e['atr_gunluk']
    eo = e['atr_ort']
    price = e['price']
    if yf_atr and eg:
        diff = yf_atr - eg
        pct  = diff / eg * 100
        diffs.append(abs(pct))
        marker = ' <<<' if abs(pct) > 25 else ''
        print(f"{sym:8s}  {price:>8.2f}  {eg:>12.4f}  {eo:>9.4f}  {yf_atr:>8.4f}  {diff:>+8.4f}  {pct:>+6.1f}%{marker}")
    else:
        print(f"{sym:8s}  {price:>8.2f}  {eg:>12.4f}  {eo:>9.4f}  {'N/A':>8s}")

if diffs:
    print('-' * 75)
    print(f"Ortalama mutlak fark: %{sum(diffs)/len(diffs):.1f}")
    print(f"Max fark: %{max(diffs):.1f}  |  Min fark: %{min(diffs):.1f}")

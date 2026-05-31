"""
price_reader.py — BIST100 anlık fiyat okuyucu
Önce Excel (MatriksIQ DDE), yoksa Yahoo Finance fallback.
"""
import time
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

EXCEL_PATH = Path(r"C:\Users\BioCSI\CLAUDE\GridTracker\Bist100 - Anlık Fiyat.xlsx")
MAX_AGE_SECONDS = 90  # Excel dosyası bu kadar saniyeden eskiyse Yahoo'ya geç


def get_price_from_excel(symbol: str) -> float | None:
    """
    Açık Excel uygulamasından DDE verisini ANINDA okur.
    Dosya kaydını beklemez — bellek içi canlı değeri alır.
    Fallback: Excel kapalıysa diskten okur (max 90 sn eski).
    """
    sym = symbol.upper().replace(".IS", "")

    # ── Yöntem 1: Açık Excel'den canlı oku (anlık) ──────────
    try:
        import win32com.client as win32
        xl = win32.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if "Anlık Fiyat" in wb.Name or "Anlik Fiyat" in wb.Name:
                ws = wb.Sheets(1)
                row = 2
                while True:
                    cell_val = ws.Cells(row, 1).Value
                    if cell_val is None:
                        break
                    if str(cell_val).upper() == sym:
                        price = ws.Cells(row, 3).Value  # C = SON
                        if price:
                            return float(price)
                        break
                    row += 1
    except:
        pass

    # ── Yöntem 2: Diskten oku (Excel kapalıysa fallback) ────
    try:
        if not EXCEL_PATH.exists():
            return None
        age = time.time() - EXCEL_PATH.stat().st_mtime
        if age > MAX_AGE_SECONDS:
            return None
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).upper() == sym:
                val = row[2]
                wb.close()
                return float(val) if val else None
        wb.close()
    except:
        pass

    return None


def get_price_from_yahoo(symbol: str) -> float | None:
    """Yahoo Finance fallback."""
    try:
        import yfinance as yf
        t = yf.Ticker(f"{symbol}.IS")
        hist = t.history(period="2d")
        if not hist.empty:
            return float(hist["Close"].iloc[-1])
    except:
        pass
    return None


def get_price(symbol: str) -> tuple[float | None, str]:
    """
    Fiyat döndürür.
    Returns: (fiyat, kaynak) — kaynak: 'excel' veya 'yahoo'
    """
    price = get_price_from_excel(symbol)
    if price and price > 0:
        return price, "excel"

    price = get_price_from_yahoo(symbol)
    if price and price > 0:
        return price, "yahoo"

    return None, "yok"


def get_all_prices() -> dict[str, float]:
    """Tüm BIST100 fiyatlarını döner. Önce açık Excel, yoksa diskten."""
    prices = {}

    # ── Açık Excel'den canlı oku ─────────────────────────────
    try:
        import win32com.client as win32
        xl = win32.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if "Anlık Fiyat" in wb.Name or "Anlik Fiyat" in wb.Name:
                ws = wb.Sheets(1)
                row = 2
                while True:
                    sym = ws.Cells(row, 1).Value
                    if sym is None:
                        break
                    price = ws.Cells(row, 3).Value
                    if price:
                        try:
                            prices[str(sym).upper()] = float(price)
                        except:
                            pass
                    row += 1
                return prices  # Başarılı, diskten okumaya gerek yok
    except:
        pass

    # ── Diskten fallback ─────────────────────────────────────
    try:
        if not EXCEL_PATH.exists():
            return prices
        age = time.time() - EXCEL_PATH.stat().st_mtime
        if age > MAX_AGE_SECONDS:
            return prices
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2]:
                try:
                    prices[str(row[0]).upper()] = float(row[2])
                except:
                    pass
        wb.close()
    except:
        pass
    return prices


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    sym = sys.argv[1].upper() if len(sys.argv) > 1 else "SASA"
    price, src = get_price(sym)
    print(f"{sym}: {price} TL  (kaynak: {src})")
    age = time.time() - EXCEL_PATH.stat().st_mtime if EXCEL_PATH.exists() else -1
    print(f"Excel yaşı: {age:.0f} sn")

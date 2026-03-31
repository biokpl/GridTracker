#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BIST Grid Bot Tracker - Otomatik Servis v2
===========================================
Her iş günü BİST kapanışından 35 dakika sonra çalışır:
  - Normal günler : 18:35
  - Arife günleri : 13:05

Masaüstündeki 1.xlsx ve 2.xlsx dosyalarını okur,
verileri doğrudan bist_tracker.html içine gömer.
(data.json da yedek olarak yazılır)

Kullanım:
  python grid_tracker_service.py          # Servis (sürekli)
  python grid_tracker_service.py --now    # Hemen çalıştır (test)
  python grid_tracker_service.py --setup  # Görev zamanlayıcıya ekle
  python grid_tracker_service.py --html   # Sadece HTML'i yeniden oluştur
"""

import os, sys, json, time, logging, argparse, re, subprocess
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl

try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    import requests


# ══════════════════════════════════════════════════════════════
#  HTML ŞABLONU  (güncelleme: bist_tracker.html bu içerikten oluşturulur)
# ══════════════════════════════════════════════════════════════
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>Grid Tracker</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
*{margin:0;padding:0;box-sizing:border-box;-webkit-tap-highlight-color:transparent;}
:root{
  --bg:#09090b;
  --surface:#111113;
  --surface2:#18181b;
  --border:#27272a;
  --border2:#3f3f46;
  --text:#fafafa;
  --text2:#a1a1aa;
  --text3:#52525b;
  --green:#22c55e;
  --red:#ef4444;
  --amber:#f59e0b;
  --blue:#3b82f6;
  --mono:'JetBrains Mono',monospace;
  --sans:'Inter',sans-serif;
  --r:10px;
}
html,body{height:100%;overflow:hidden;}
body{background:var(--bg);color:var(--text);font-family:var(--sans);display:flex;flex-direction:column;font-size:14px;}

/* ── TOPBAR ── */
.topbar{
  height:52px;min-height:52px;
  display:flex;align-items:center;justify-content:space-between;
  padding:0 16px;
  border-bottom:1px solid var(--border);
  background:var(--surface);
}
.brand{display:flex;align-items:center;gap:10px;}
.brand-mark{
  width:28px;height:28px;border-radius:7px;
  background:linear-gradient(135deg,#3b82f6,#1d4ed8);
  display:flex;align-items:center;justify-content:center;
  font-size:13px;font-weight:700;color:#fff;letter-spacing:-0.5px;
  font-family:var(--mono);
}
.brand-name{font-size:13px;font-weight:600;letter-spacing:.5px;color:var(--text);}
.brand-name span{color:var(--text3);font-weight:400;}
.topbar-right{display:flex;align-items:center;gap:10px;}
.market-pill{
  display:flex;align-items:center;gap:5px;
  padding:4px 10px;border-radius:20px;
  background:var(--surface2);border:1px solid var(--border);
  font-family:var(--mono);font-size:11px;color:var(--text2);
}
.dot{width:6px;height:6px;border-radius:50%;background:var(--green);animation:blink 2s infinite;}
.dot.off{background:var(--red);animation:none;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.3}}
.sync-time{font-family:var(--mono);font-size:10px;color:var(--text3);}

/* ── NAV ── */
.nav{
  display:flex;gap:2px;padding:6px;
  border-bottom:1px solid var(--border);
  background:var(--surface);
  overflow-x:auto;scrollbar-width:none;
}
.nav::-webkit-scrollbar{display:none;}
.nav-btn{
  flex:1;min-width:64px;
  padding:7px 12px;border-radius:6px;border:none;
  background:transparent;color:var(--text3);
  font-family:var(--sans);font-size:12px;font-weight:500;
  cursor:pointer;transition:all .15s;white-space:nowrap;
  display:flex;align-items:center;justify-content:center;gap:5px;
}
.nav-btn:hover{color:var(--text2);background:var(--surface2);}
.nav-btn.on{background:var(--surface2);color:var(--text);border:1px solid var(--border);}
.nav-icon{font-size:13px;}

/* ── SCROLL AREA ── */
.content{flex:1;overflow-y:auto;overflow-x:hidden;scrollbar-width:thin;scrollbar-color:rgba(59,130,246,.25) transparent;}
.content::-webkit-scrollbar{width:4px;}
.content::-webkit-scrollbar-track{background:transparent;}
.content::-webkit-scrollbar-thumb{background:linear-gradient(to bottom,rgba(59,130,246,.15),rgba(59,130,246,.4),rgba(59,130,246,.15));border-radius:4px;transition:background .2s;}
.content::-webkit-scrollbar-thumb:hover{background:linear-gradient(to bottom,rgba(59,130,246,.3),rgba(59,130,246,.7),rgba(59,130,246,.3));}
.page{display:none;padding:14px;max-width:600px;margin:0 auto;}
.page.on{display:block;}

/* ── ALERT ── */
.alert{
  padding:10px 14px;border-radius:var(--r);margin-bottom:12px;
  font-size:12px;display:flex;align-items:center;gap:8px;
  background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.2);color:var(--text2);
}
.alert.warn{background:rgba(245,158,11,.08);border-color:rgba(245,158,11,.2);}
.alert.err{background:rgba(239,68,68,.08);border-color:rgba(239,68,68,.2);color:#fca5a5;}
.alert.ok{background:rgba(34,197,94,.08);border-color:rgba(34,197,94,.2);color:#86efac;}
.alert-icon{font-size:14px;flex-shrink:0;}

/* ── PORTFÖY KARTI ── */
.portfolio-card{
  position:relative;
  background:linear-gradient(160deg,#0f172a 0%,#111827 50%,#0f1629 100%);
  border:1px solid #1e3a5f;
  border-radius:16px;
  padding:28px 20px 22px;
  margin-bottom:12px;
  text-align:center;
  overflow:hidden;
}
.portfolio-card::before{
  content:'';position:absolute;inset:0;
  background-image:
    linear-gradient(rgba(59,130,246,.04) 1px,transparent 1px),
    linear-gradient(90deg,rgba(59,130,246,.04) 1px,transparent 1px);
  background-size:24px 24px;pointer-events:none;
}
.portfolio-card::after{
  content:'';position:absolute;top:0;left:10%;right:10%;height:1px;
  background:linear-gradient(90deg,transparent,#3b82f6,#93c5fd,#3b82f6,transparent);
}
.pf-glow{
  position:absolute;top:-60px;left:50%;transform:translateX(-50%);
  width:220px;height:120px;
  background:radial-gradient(ellipse,rgba(59,130,246,.2) 0%,transparent 70%);
  pointer-events:none;
}
.pf-label{
  position:relative;
  font-size:10px;font-weight:700;letter-spacing:2.8px;
  text-transform:uppercase;color:#8b9ab5;margin-bottom:14px;
}
.pf-value{
  position:relative;
  font-family:var(--mono);font-size:44px;font-weight:600;
  letter-spacing:-2px;line-height:1;margin-bottom:6px;
  background:linear-gradient(180deg,#e2e8f0 0%,#94a3b8 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  background-clip:text;
}
.pf-currency{font-size:22px;font-weight:400;letter-spacing:0;}
.pf-divider{
  position:relative;height:1px;
  background:linear-gradient(90deg,transparent,#1e3a5f 30%,#1e3a5f 70%,transparent);
  margin:16px 0;
}
.pf-bottom{
  position:relative;
  display:flex;align-items:center;justify-content:center;gap:28px;
}
.pf-stat{text-align:center;}
.pf-stat-label{font-size:9px;letter-spacing:1.8px;text-transform:uppercase;color:#8b9ab5;margin-bottom:5px;font-weight:600;}
.pf-stat-val{font-family:var(--mono);font-size:15px;font-weight:500;color:#94a3b8;}
.pf-stat-val.pos{color:#4ade80;}
.pf-stat-val.pos::before{content:'▲ ';}
.pf-stat-val.neg{color:#f87171;}
.pf-stat-val.neg::before{content:'▼ ';}
.pf-sep{width:1px;height:30px;background:#1e3a5f;}

/* ── KPI STRIP ── */
.kpi-strip{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px;}
.kpi-strip.three{grid-template-columns:1fr 1fr 1fr;}
.kpi{
  background:var(--surface);border:1px solid var(--border);
  border-radius:var(--r);padding:14px 10px 12px;
  position:relative;overflow:hidden;
  text-align:center;
  transition:border-color .2s,box-shadow .2s;
}
.kpi:hover{border-color:var(--border2);box-shadow:0 0 0 1px rgba(255,255,255,.03);}
/* bottom accent line */
.kpi::before{
  content:'';position:absolute;bottom:0;left:0;right:0;height:2px;
  background:var(--kpi-accent,linear-gradient(90deg,transparent,var(--blue),transparent));
  opacity:.65;
}
/* top shimmer line */
.kpi::after{
  content:'';position:absolute;top:0;left:10%;right:10%;height:1px;
  background:var(--kpi-accent,linear-gradient(90deg,transparent,var(--blue),transparent));
  opacity:.2;
}
.kpi.g{--kpi-accent:linear-gradient(90deg,transparent,var(--green),transparent);background:linear-gradient(180deg,rgba(34,197,94,.03) 0%,var(--surface) 55%);}
.kpi.a{--kpi-accent:linear-gradient(90deg,transparent,var(--amber),transparent);background:linear-gradient(180deg,rgba(245,158,11,.03) 0%,var(--surface) 55%);}
.kpi.r{--kpi-accent:linear-gradient(90deg,transparent,var(--red),transparent);background:linear-gradient(180deg,rgba(239,68,68,.03) 0%,var(--surface) 55%);}
.kpi.pu{--kpi-accent:linear-gradient(90deg,transparent,#a78bfa,transparent);background:linear-gradient(180deg,rgba(167,139,250,.03) 0%,var(--surface) 55%);}
.kpi-icon{font-size:11px;display:inline;line-height:1;margin-right:4px;text-transform:none;letter-spacing:0;}
.kpi-label{font-size:9px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#8b9ab5;margin-bottom:7px;display:flex;align-items:center;justify-content:center;gap:0;}
.kpi-val{font-family:var(--mono);font-size:18px;font-weight:500;color:var(--text);line-height:1;}
.kpi-val.pos{color:#4ade80;}
.kpi-val.neg{color:var(--red);}
.kpi-val.amb{color:#fbbf24;}
.kpi-val.blue{color:#60a5fa;}
.kpi-val.pu{color:#c4b5fd;}
.kpi-sub{font-size:10px;color:#6b7a92;margin-top:5px;}
.kpi-val.kpi-muted{color:#7b8fa8;}
.kpi-wide{grid-column:1/-1;}

/* ── SECTION ── */
.section{margin-bottom:16px;}
.section-head{
  display:flex;align-items:center;justify-content:space-between;
  margin-bottom:10px;
}
.section-title{
  font-size:11px;font-weight:700;letter-spacing:1.4px;text-transform:uppercase;
  color:#a1a1aa;
  display:flex;align-items:center;gap:8px;
}
.section-accent{
  display:inline-block;width:2px;height:13px;border-radius:2px;
  background:linear-gradient(to bottom,var(--blue),rgba(59,130,246,.15));
  flex-shrink:0;
}
.section-icon{font-size:12px;opacity:.75;}
.section-badge{
  font-family:var(--mono);font-size:10px;
  padding:3px 8px;border-radius:5px;
  background:rgba(59,130,246,.08);color:#93c5fd;border:1px solid rgba(59,130,246,.2);
  font-weight:600;letter-spacing:.5px;
}

/* ── CARD ── */
.card{
  background:var(--surface);border:1px solid var(--border);
  border-radius:var(--r);overflow:hidden;margin-bottom:8px;
}
.card-head{
  padding:11px 14px;border-bottom:1px solid var(--border);
  display:flex;align-items:center;justify-content:space-between;
}
.card-title{font-size:12px;font-weight:600;color:var(--text2);}
.card-body{padding:14px;}

/* ── TABLE ── */
.tbl-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;font-family:var(--mono);font-size:11px;}
th{
  padding:9px 10px;text-align:left;
  background:var(--surface2);
  color:#8b9ab5;font-size:9px;letter-spacing:1.4px;text-transform:uppercase;
  border-bottom:1px solid var(--border);white-space:nowrap;font-weight:700;
}
td{
  padding:9px 10px;
  border-bottom:1px solid rgba(39,39,42,.6);
  white-space:nowrap;color:var(--text2);
}
tr:last-child td{border-bottom:none;}
tr:hover td{background:rgba(255,255,255,.02);}
.td-sym{color:#7b92b5;font-weight:600;}
.td-pos{color:var(--green);}
.td-neg{color:var(--red);}
.td-amb{color:var(--amber);}
.td-dim{color:var(--text3);}
.htbl th,.htbl td{text-align:center;}
.td-sell{color:#f87171;}
.th-sell,.th-netkar{display:inline-block;position:relative;padding-bottom:4px;}
.th-sell::after{
  content:'';position:absolute;bottom:0;left:-40%;width:180%;height:1px;border-radius:1px;
  background:linear-gradient(90deg,rgba(248,113,113,0),rgba(248,113,113,.72) 35%,rgba(252,165,165,.9) 50%,rgba(248,113,113,.72) 65%,rgba(248,113,113,0));
}
.th-netkar::after{
  content:'';position:absolute;bottom:0;left:-40%;width:180%;height:1px;border-radius:1px;
  background:linear-gradient(90deg,rgba(34,197,94,0),rgba(34,197,94,.65) 35%,rgba(74,222,128,.85) 50%,rgba(34,197,94,.65) 65%,rgba(34,197,94,0));
}
.pill{
  display:inline-block;padding:3px 9px;border-radius:5px;
  font-size:9px;font-weight:700;letter-spacing:1px;text-transform:uppercase;
}
.pill-b{background:rgba(34,197,94,.12);color:#4ade80;border:1px solid rgba(34,197,94,.22);}
.pill-s{background:rgba(248,113,113,.12);color:#f87171;border:1px solid rgba(248,113,113,.22);}
.pill-m{background:rgba(59,130,246,.12);color:#93c5fd;border:1px solid rgba(59,130,246,.22);}
/* trades filter bar */
.t-filter{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap;align-items:center;
  background:var(--surface2);border:1px solid var(--border);border-radius:10px;padding:10px 12px;}
.t-filter-group{display:flex;flex-direction:column;gap:3px;flex:1;min-width:100px;}
.t-filter-label{font-size:9px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#4e6080;}
.t-filter select{
  background:var(--surface);border:1px solid var(--border);border-radius:6px;
  padding:7px 10px;color:#a1b0c8;font-family:var(--mono);font-size:11px;
  outline:none;cursor:pointer;transition:border-color .15s;width:100%;
}
.t-filter select:focus{border-color:rgba(59,130,246,.4);}
.t-filter select option{background:#1a1f2e;}
/* trades table */
.ttbl th,.ttbl td{text-align:center;}
.td-time{color:#4e6080;font-size:10px;font-family:var(--mono);}
.td-lot{color:#c4cfe0;font-family:var(--mono);}
.td-price{color:#8fa8c8;font-family:var(--mono);}
.td-amount{color:#c4cfe0;font-weight:600;}

/* ── INPUT ── */
.field{display:flex;flex-direction:column;gap:4px;margin-bottom:10px;}
.field-label{font-size:10px;font-weight:500;letter-spacing:.8px;text-transform:uppercase;color:var(--text3);}
.field-input{
  background:var(--surface2);border:1px solid var(--border);border-radius:7px;
  padding:9px 12px;color:var(--text);font-family:var(--mono);font-size:13px;
  outline:none;transition:border-color .15s;width:100%;
}
.field-input:focus{border-color:var(--blue);}
.field-input::placeholder{color:var(--text3);}
.field-input[type=number]{-moz-appearance:textfield;}
.field-input[type=number]::-webkit-inner-spin-button{-webkit-appearance:none;}
select.field-input{cursor:pointer;}
select.field-input option{background:var(--surface2);}

.row-fields{display:grid;grid-template-columns:1fr 1fr;gap:8px;}
.row-fields.three{grid-template-columns:1fr 1fr 1fr;}

/* ── BUTTON ── */
.btn{
  padding:9px 16px;border-radius:7px;border:none;
  font-family:var(--sans);font-size:12px;font-weight:600;
  cursor:pointer;transition:all .15s;display:inline-flex;align-items:center;gap:6px;
}
.btn-primary{background:var(--blue);color:#fff;}
.btn-primary:hover{background:#2563eb;}
.btn-ghost{background:transparent;border:1px solid var(--border);color:var(--text2);}
.btn-ghost:hover{border-color:var(--border2);color:var(--text);}
.btn-danger{background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);color:var(--red);}
.btn-sm{padding:6px 12px;font-size:11px;}
.btn-xs{padding:4px 8px;font-size:10px;}
.btn-full{width:100%;justify-content:center;}
.btn-row{display:flex;gap:8px;margin-top:10px;flex-wrap:wrap;}

/* ── LICENSE BANNER ── */
.lic-banner{margin-bottom:14px;margin-top:6px;padding-top:14px;}
.lic-track{height:2px;background:rgba(255,255,255,.05);border-radius:2px;overflow:visible;margin-bottom:5px;}
.lic-glow{
  height:2px;border-radius:2px;min-width:6px;
  transition:width 1.2s cubic-bezier(.4,0,.2,1);
  background:linear-gradient(90deg,rgba(251,191,36,0),rgba(180,130,40,.35) 15%,rgba(251,191,36,.78) 42%,rgba(253,211,77,.92) 50%,rgba(251,191,36,.78) 58%,rgba(180,130,40,.35) 85%,rgba(251,191,36,0));
  filter:drop-shadow(0 0 4px rgba(251,191,36,.7)) drop-shadow(0 0 10px rgba(251,191,36,.3));
}
.lic-meta{display:flex;justify-content:space-between;align-items:center;}
.lic-meta-txt{font-size:10px;color:#6b7a92;opacity:.85;}
.lic-meta-pct{font-family:var(--mono);font-size:10px;color:#6b7a92;opacity:.85;}
/* hidden compat elements */
.lic-icon,.lic-body,.lic-title,.lic-sub,.lic-right,.lic-right-lbl,.lic-right-val,.lic-pct,.prog-wrap{display:none!important;}

/* ── SETTINGS CENTERED HEADER ── */
.s-head-c{display:flex;align-items:center;gap:10px;margin-bottom:12px;position:relative;}
.s-rule{flex:1;height:1px;background:linear-gradient(to right,transparent,rgba(59,130,246,.2),transparent);}
.s-title-c{font-size:11px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#a1a1aa;white-space:nowrap;display:flex;align-items:center;gap:0;}

/* ── MONTH NAV ── */
.month-nav{display:flex;align-items:center;gap:4px;margin-bottom:16px;}
.month-btn{
  background:var(--surface2);border:1px solid var(--border);border-radius:7px;
  color:#4e6080;padding:6px 12px;cursor:pointer;font-size:14px;
  transition:all .15s;flex-shrink:0;
}
.month-btn:hover{border-color:rgba(59,130,246,.3);color:#93c5fd;}
.month-label{flex:1;text-align:center;}
.month-label-m{font-size:13px;font-weight:600;color:#8fa8c8;letter-spacing:.3px;display:block;}
.month-label-y{font-size:10px;font-weight:400;color:#3d5070;letter-spacing:1.5px;display:block;margin-top:2px;}

/* ── EMPTY ── */
.empty{text-align:center;padding:32px 20px;color:var(--text3);}
.empty-icon{font-size:28px;margin-bottom:8px;opacity:.5;}
.empty-text{font-size:12px;}

/* ── DIVIDER ── */
.div{height:1px;background:var(--border);margin:12px 0;}

/* ── COST ITEM ── */
.cost-row{display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid rgba(39,39,42,.5);}
.cost-row:last-child{border-bottom:none;}
.cost-name{flex:1;font-size:11px;color:#7b8fa8;letter-spacing:.2px;}
.cost-val{font-family:var(--mono);font-size:12px;font-weight:500;color:var(--amber);}
.cost-edit{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:5px;
  padding:4px 7px;color:#a1b0c8;font-family:var(--mono);font-size:11px;width:78px;outline:none;text-align:right;
  -moz-appearance:textfield;}
.cost-edit::-webkit-inner-spin-button,.cost-edit::-webkit-outer-spin-button{-webkit-appearance:none;margin:0;}
.cost-edit:focus{border-color:rgba(59,130,246,.35);}
.cost-unit{font-size:10px;color:#4e6080;}
/* cost summary strip */
.cost-summary{display:flex;justify-content:space-around;align-items:center;
  padding:8px 4px 2px;border-top:1px solid rgba(255,255,255,.05);margin-top:6px;}
.cost-sum-item{text-align:center;}
.cost-sum-label{font-size:8px;font-weight:700;letter-spacing:1.2px;text-transform:uppercase;color:#4e6080;margin-bottom:3px;}
.cost-sum-val{font-family:var(--mono);font-size:13px;font-weight:600;}

/* ── TOAST ── */
.toasts{position:fixed;bottom:16px;right:16px;z-index:999;display:flex;flex-direction:column;gap:6px;}
.toast{
  background:var(--surface2);border:1px solid var(--border);
  padding:10px 14px;border-radius:8px;font-size:12px;
  max-width:280px;display:flex;align-items:center;gap:8px;
  animation:tin .25s ease;
}
.toast.ok{border-color:rgba(34,197,94,.4);color:#86efac;}
.toast.err{border-color:rgba(239,68,68,.4);color:#fca5a5;}
.toast.inf{border-color:rgba(59,130,246,.4);color:#93c5fd;}
@keyframes tin{from{transform:translateX(20px);opacity:0}to{transform:translateX(0);opacity:1}}

/* ── INNER TABS ── */
.itabs{display:flex;gap:4px;margin-bottom:12px;background:var(--surface2);border-radius:8px;padding:3px;}
.itab{
  flex:1;padding:6px;border-radius:5px;border:none;background:transparent;
  color:#4e6080;font-size:11px;font-weight:600;cursor:pointer;transition:all .15s;letter-spacing:.4px;
}
.itab.on{background:rgba(59,130,246,.12);color:#93c5fd;border:1px solid rgba(59,130,246,.25);}
.ipage{display:none;} .ipage.on{display:block;}

/* ── RESPONSIVE ── */
@media(min-width:600px){
  .page{padding:20px;}
  .kpi-val{font-size:22px;}
}
@media(max-width:380px){
  .kpi-val{font-size:15px;}
  .kpi-strip.three .kpi-val{font-size:13px;}
}
</style>
</head>
<body>

<!-- TOPBAR -->
<div class="topbar">
  <div class="brand">
    <div class="brand-mark">G</div>
    <div class="brand-name">Grid<span>Tracker</span></div>
  </div>
  <div class="topbar-right">
    <div class="sync-time" id="syncTime">—</div>
    <div class="market-pill">
      <div class="dot" id="mDot"></div>
      <span id="mTxt">—</span>
    </div>
  </div>
</div>

<!-- NAV -->
<div class="nav">
  <button class="nav-btn on" onclick="go('pg-home',this)"><span class="nav-icon">◆</span>Özet</button>
  <button class="nav-btn" onclick="go('pg-trades',this)"><span class="nav-icon">⇅</span>İşlemler</button>
  <button class="nav-btn" onclick="go('pg-monthly',this)"><span class="nav-icon">▦</span>Aylık</button>
  <button class="nav-btn" onclick="go('pg-overall',this)"><span class="nav-icon">◎</span>Overall</button>
  <button class="nav-btn" onclick="go('pg-settings',this)"><span class="nav-icon">⚙</span>Ayarlar</button>
</div>

<!-- CONTENT -->
<div class="content">

<!-- ══ ÖZET ══ -->
<div class="page on" id="pg-home">

  <div class="alert" id="syncAlert" style="display:none">
    <span class="alert-icon">ℹ</span>
    <span id="syncMsg">—</span>
  </div>

  <!-- Overall — ana kart -->
  <div class="portfolio-card">
    <div class="pf-glow"></div>
    <div class="pf-label">Portföy Değeri</div>
    <div class="pf-value" id="h-overall">—</div>
    <div class="pf-divider"></div>
    <div class="pf-bottom">
      <div class="pf-stat">
        <div class="pf-stat-label">Önceki Güne Göre</div>
        <div class="pf-stat-val" id="h-overall-diff">—</div>
      </div>
      <div class="pf-sep"></div>
      <div class="pf-stat">
        <div class="pf-stat-label">Tarih</div>
        <div class="pf-stat-val" id="h-overall-date">—</div>
      </div>
    </div>
  </div>

  <!-- Günlük P&L -->
  <div class="kpi-strip three" style="margin-bottom:10px">
    <div class="kpi g">
      <div class="kpi-label"><span class="kpi-icon">📈</span>Brüt Kar</div>
      <div class="kpi-val pos" id="h-gross">—</div>
      <div class="kpi-sub">₺ bugün</div>
    </div>
    <div class="kpi a">
      <div class="kpi-label"><span class="kpi-icon">💸</span>Komisyon</div>
      <div class="kpi-val amb" id="h-comm">—</div>
      <div class="kpi-sub">₺ bugün</div>
    </div>
    <div class="kpi">
      <div class="kpi-label"><span class="kpi-icon">💎</span>Net Kar</div>
      <div class="kpi-val" id="h-net">—</div>
      <div class="kpi-sub">₺ bugün</div>
    </div>
  </div>

  <!-- İşlem özeti -->
  <div class="kpi-strip" style="margin-bottom:14px">
    <div class="kpi pu">
      <div class="kpi-label"><span class="kpi-icon">⚡</span>Toplam İşlem</div>
      <div class="kpi-val pu" id="h-trades">—</div>
      <div class="kpi-sub" id="h-trades-sub">— alış · — satış</div>
    </div>
    <div class="kpi">
      <div class="kpi-label"><span class="kpi-icon">📅</span>Aylık Net Kar</div>
      <div class="kpi-val" id="h-mnet">—</div>
      <div class="kpi-sub" id="h-mdays">—</div>
    </div>
  </div>

  <!-- Lisans banner -->
  <div class="lic-banner lic-no" id="licBanner" style="display:none">
    <div class="lic-track"><div class="lic-glow" id="licBar" style="width:0%"></div></div>
    <div class="lic-meta">
      <span class="lic-meta-txt" id="licSub">—</span>
      <span class="lic-meta-pct" id="licPct">%0</span>
    </div>
    <!-- compat -->
    <span id="licIcon" style="display:none"></span>
    <span id="licTitle" style="display:none"></span>
    <span id="licTarget" style="display:none"></span>
  </div>

  <!-- Hisse özeti -->
  <div class="section">
    <div class="section-head">
      <div class="section-title"><span class="section-accent"></span><span class="section-icon">📊</span>Hisse Özeti</div>
      <div style="display:flex;align-items:center;gap:8px">
        <div class="itabs" style="margin:0">
          <button class="itab sym-tab" onclick="setSymMode('all',this)">Tümü</button>
          <button class="itab on sym-tab" onclick="setSymMode('bot',this)">GridBot</button>
        </div>
        <div class="section-badge" id="h-sym-count">—</div>
      </div>
    </div>
    <div class="card">
      <div class="tbl-wrap">
        <table class="htbl">
          <thead><tr><th>Hisse</th><th>Alış</th><th><span class="th-sell">Satış</span></th><th>Ort.Alış</th><th>Ort.Satış</th><th><span class="th-netkar">Net Kar</span></th></tr></thead>
          <tbody id="h-sym-tbl"><tr><td colspan="6"><div class="empty"><div class="empty-icon">📊</div><div class="empty-text">Veri bekleniyor</div></div></td></tr></tbody>
        </table>
      </div>
    </div>
  </div>

</div><!-- /pg-home -->

<!-- ══ İŞLEMLER ══ -->
<div class="page" id="pg-trades">

  <div class="s-head-c" style="margin-bottom:12px"><div class="s-rule"></div><div class="section-title s-title-c">İşlem Geçmişi</div><div class="s-rule"></div></div>

  <div class="t-filter">
    <div class="t-filter-group">
      <div class="t-filter-label">🏷 Hisse</div>
      <select id="fSym" onchange="renderTrades()">
        <option value="">Tüm Hisseler</option>
      </select>
    </div>
    <div class="t-filter-group">
      <div class="t-filter-label">⇅ İşlem Tipi</div>
      <select id="fType" onchange="renderTrades()">
        <option value="">Alış &amp; Satış</option>
        <option>Alış</option>
        <option>Satış</option>
      </select>
    </div>
    <div style="display:flex;align-items:flex-end;padding-bottom:1px">
      <button class="btn btn-ghost btn-sm" onclick="exportCSV()" style="white-space:nowrap">↓ CSV</button>
    </div>
  </div>

  <div class="card">
    <div class="tbl-wrap">
      <table class="ttbl">
        <thead><tr><th>Saat</th><th>Hisse</th><th>İşlem</th><th>Lot</th><th>Fiyat</th><th>Tutar</th><th>Komisyon</th></tr></thead>
        <tbody id="trades-tbl"><tr><td colspan="7"><div class="empty"><div class="empty-icon">↕</div><div class="empty-text">Veri yok</div></div></td></tr></tbody>
      </table>
    </div>
  </div>

</div><!-- /pg-trades -->

<!-- ══ AYLIK ══ -->
<div class="page" id="pg-monthly">

  <div class="month-nav">
    <button class="month-btn" onclick="changeMonth(-1)">‹</button>
    <div class="month-label" id="monthLbl">—</div>
    <button class="month-btn" onclick="changeMonth(1)">›</button>
  </div>

  <div class="kpi-strip three" style="margin-bottom:12px">
    <div class="kpi g">
      <div class="kpi-label">Brüt Kar</div>
      <div class="kpi-val" id="m-gross">—</div>
    </div>
    <div class="kpi a">
      <div class="kpi-label">Komisyon</div>
      <div class="kpi-val amb" id="m-comm">—</div>
    </div>
    <div class="kpi">
      <div class="kpi-label">Net Kar</div>
      <div class="kpi-val" id="m-net">—</div>
    </div>
  </div>

  <div class="kpi-strip" style="margin-bottom:16px">
    <div class="kpi" style="padding:0;display:flex;overflow:hidden">
      <div style="flex:1;padding:14px 10px 12px;text-align:center">
        <div class="kpi-label">≈ Ort. Günlük Kar</div>
        <div class="kpi-val kpi-muted" id="m-avg">—</div>
      </div>
      <div style="width:1px;background:rgba(255,255,255,.06);margin:10px 0"></div>
      <div style="flex:1;padding:14px 10px 12px;text-align:center">
        <div class="kpi-label">≈ Ort. Aylık Kar</div>
        <div class="kpi-val kpi-muted" id="m-mavg">—</div>
      </div>
    </div>
    <div class="kpi">
      <div class="kpi-label">Overall Değişimi</div>
      <div class="kpi-val" id="m-odiff">—</div>
      <div class="kpi-sub" id="m-days">— işlem günü</div>
    </div>
  </div>

  <div class="s-head-c"><div class="s-rule"></div><div class="section-title s-title-c">Günlük Detay</div><div class="s-rule"></div></div>
  <div class="card">
    <div class="tbl-wrap">
      <table class="ttbl">
        <thead><tr><th>Tarih</th><th>İşlem</th><th>Brüt Kar</th><th>Komisyon</th><th><span class="th-netkar">Net Kar</span></th><th>Overall</th></tr></thead>
        <tbody id="m-tbl"><tr><td colspan="6"><div class="empty"><div class="empty-icon">📅</div><div class="empty-text">Bu ay veri yok</div></div></td></tr></tbody>
      </table>
    </div>
  </div>

</div><!-- /pg-monthly -->

<!-- ══ OVERALL ══ -->
<div class="page" id="pg-overall">

  <div class="s-head-c" style="margin-bottom:10px;position:relative">
    <div class="s-rule"></div><div class="section-title s-title-c">Portföy Geçmişi</div><div class="s-rule"></div>
    <button class="btn btn-ghost btn-xs" style="position:absolute;right:0;top:50%;transform:translateY(-50%);font-size:10px" onclick="toggleOForm()" id="oFormBtn">+ Kayıt</button>
  </div>

  <!-- kayıt formu -->
  <div id="oForm" style="display:none;margin-bottom:10px">
    <div class="card">
      <div class="card-body">
        <div class="row-fields">
          <div class="field"><div class="field-label">📅 Tarih</div><input class="field-input" id="oDate" type="date"></div>
          <div class="field"><div class="field-label">₺ Portföy Değeri</div><input class="field-input" id="oAmt" type="number" placeholder="0"></div>
        </div>
        <div class="field"><div class="field-label">💬 Not</div><input class="field-input" id="oNote" placeholder="Opsiyonel"></div>
        <div style="display:flex;gap:8px;margin-top:6px">
          <button class="btn btn-primary btn-sm" style="flex:1" onclick="saveOverall()">Kaydet</button>
          <button class="btn btn-ghost btn-sm" onclick="toggleOForm()">İptal</button>
        </div>
      </div>
    </div>
  </div>

  <div class="card">
    <div class="tbl-wrap">
      <table class="ttbl">
        <thead><tr><th>Tarih</th><th>Portföy Değeri</th><th>Değişim</th><th>Not</th><th style="width:32px"></th></tr></thead>
        <tbody id="o-tbl"><tr><td colspan="5"><div class="empty"><div class="empty-icon">📊</div><div class="empty-text">Henüz kayıt yok</div></div></td></tr></tbody>
      </table>
    </div>
  </div>

</div><!-- /pg-overall -->

<!-- ══ AYARLAR ══ -->
<div class="page" id="pg-settings">

  <!-- Genel -->
  <div class="section">
    <div class="s-head-c"><div class="s-rule"></div><div class="section-title s-title-c">Genel</div><div class="s-rule"></div></div>
    <div class="card">
      <div class="card-body" style="padding:10px 12px;text-align:center">
        <div style="display:inline-flex;align-items:center;gap:8px;margin-bottom:5px">
          <span style="font-size:10px;color:#7b8fa8;font-weight:600;letter-spacing:.3px">Komisyon Oranı</span>
          <span style="font-family:var(--mono);font-size:11px;color:#4ade80;background:rgba(34,197,94,.07);border:1px solid rgba(34,197,94,.15);padding:3px 10px;border-radius:20px;font-weight:600;letter-spacing:.5px">%0,01</span>
        </div>
        <div style="font-size:9px;color:#4e6080;font-style:italic">Her alış ve satış için ayrı uygulanır</div>
      </div>
    </div>
    <input type="hidden" id="sComm" value="1">
    <input type="hidden" id="sGrid" value="0">
    <input type="hidden" id="sTol" value="0">
  </div>

  <!-- GridBot Hisseleri -->
  <div class="section">
    <div class="s-head-c"><div class="s-rule"></div><div class="section-title s-title-c">GridBot Hisseleri</div><div class="s-rule"></div></div>
    <div class="card">
      <div class="card-body" style="padding:10px 12px">
        <div style="font-size:10px;color:#4e6080;margin-bottom:10px;text-align:center;font-style:italic">
          Hisse Özeti'nde <strong style="color:#7aa6e0;font-style:normal">GridBot</strong> görünümü için filtrelenecek hisseler.
        </div>
        <div style="display:flex;gap:6px;margin-bottom:8px">
          <input class="field-input" id="botSymInput" placeholder="Hisse kodu" style="text-transform:uppercase;flex:1;padding:7px 10px;font-size:11px"
            onkeydown="if(event.key==='Enter') addBotSym()">
          <button class="btn btn-ghost btn-sm" style="font-size:10px;padding:0 12px" onclick="addBotSym()">+ Ekle</button>
        </div>
        <div id="botSymList"></div>
      </div>
    </div>
  </div>

  <!-- Lisans Giderleri -->
  <div class="section">
    <div class="s-head-c" style="position:relative">
      <div class="s-rule"></div><div class="section-title s-title-c">Lisans Giderleri</div><div class="s-rule"></div>
      <button class="btn btn-ghost btn-xs" style="position:absolute;right:0;top:50%;transform:translateY(-50%)" onclick="addCost()">+ Ekle</button>
    </div>
    <div class="card">
      <div class="card-body" style="padding:10px 12px">
        <div style="font-size:10px;color:#4e6080;margin-bottom:8px;text-align:center;font-style:italic">
          Aylık komisyon, toplam giderin 3 katını aşarsa MatriksIQ lisansı ücretsiz olur.
        </div>
        <div id="costList"></div>
        <div class="cost-summary">
          <div class="cost-sum-item">
            <div class="cost-sum-label">Aylık Toplam</div>
            <div class="cost-sum-val" style="color:var(--amber)" id="sTotalCost">—</div>
          </div>
          <div style="width:1px;height:28px;background:rgba(255,255,255,.06)"></div>
          <div class="cost-sum-item">
            <div class="cost-sum-label">Komisyon Hedefi</div>
            <div class="cost-sum-val" style="color:#60a5fa" id="sTarget">—</div>
          </div>
          <div style="width:1px;height:28px;background:rgba(255,255,255,.06)"></div>
          <div class="cost-sum-item">
            <div class="cost-sum-label">Bu Ay Komisyon</div>
            <div class="cost-sum-val" style="color:#4ade80" id="sEarned">—</div>
          </div>
        </div>
        <button class="btn btn-ghost btn-sm btn-full" style="margin-top:8px;font-size:10px" onclick="saveCosts()">Kaydet</button>
      </div>
    </div>
  </div>

  <!-- Sabah Otomasyonu -->
  <div class="section">
    <div class="s-head-c"><div class="s-rule"></div><div class="section-title s-title-c">Sabah Otomasyonu</div><div class="s-rule"></div></div>
    <div class="card">
      <div class="card-body" style="padding:10px 12px">
        <div id="autoStatus" class="alert" style="display:none;margin-bottom:8px"></div>
        <div class="row-fields" style="gap:8px">
          <div class="field" style="margin-bottom:0">
            <div class="field-label" style="font-size:8px;color:#4e6080;margin-bottom:4px">Çalışma Saati</div>
            <input class="field-input" id="autoTime" type="time" value="09:15" style="padding:7px 10px;font-size:12px">
            <div style="font-size:9px;color:#3d5070;margin-top:3px;font-style:italic">Pzt – Cum, tatiller hariç</div>
          </div>
          <div class="field" style="margin-bottom:0">
            <div class="field-label" style="font-size:8px;color:#4e6080;margin-bottom:4px">BIOS Uyandırma</div>
            <div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:6px;padding:7px 10px">
              <span style="font-family:var(--mono);font-size:12px;color:var(--amber)" id="autoWake">—</span>
            </div>
            <div style="font-size:9px;color:#3d5070;margin-top:3px;font-style:italic">PC'yi 5 dk önce aç</div>
          </div>
        </div>
        <button class="btn btn-ghost btn-sm btn-full" style="margin-top:8px;font-size:10px" onclick="saveAutoSettings()">Kaydet &amp; Görevi Güncelle</button>
        <div id="biosReminder" style="display:none;margin-top:8px;padding:8px 12px;border-radius:7px;background:rgba(245,158,11,.08);border:1px solid rgba(245,158,11,.2);">
          <div style="font-size:10px;color:var(--amber);font-weight:600;margin-bottom:3px">⚠ BIOS Ayarını Güncelle</div>
          <div style="font-size:10px;color:#7b8fa8;line-height:1.5">BIOS → Power Management → <strong style="color:#94a3b8">RTC Alarm</strong> saatini <span id="biosWakeDisplay" style="font-family:var(--mono);color:var(--amber)"></span> olarak ayarla.</div>
        </div>
      </div>
    </div>
  </div>

  <!-- Veri Yönetimi -->
  <div class="section">
    <div class="s-head-c"><div class="s-rule"></div><div class="section-title s-title-c">Veri Yönetimi</div><div class="s-rule"></div></div>
    <div class="card">
      <div class="card-body" style="padding:10px 12px">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px">
          <button class="btn btn-ghost btn-sm" style="font-size:10px;flex-direction:column;gap:2px;padding:8px 4px" onclick="exportJSON()">
            <span style="font-size:13px">⬇</span><span>Yedek Al</span>
          </button>
          <button class="btn btn-ghost btn-sm" style="font-size:10px;flex-direction:column;gap:2px;padding:8px 4px" onclick="document.getElementById('importFile').click()">
            <span style="font-size:13px">⬆</span><span>Yedek Yükle</span>
          </button>
          <button class="btn btn-danger btn-sm" style="font-size:10px;flex-direction:column;gap:2px;padding:8px 4px" onclick="clearAll()">
            <span style="font-size:13px">✕</span><span>Sıfırla</span>
          </button>
        </div>
        <input type="file" id="importFile" accept=".json" style="display:none" onchange="importJSON(event)">
      </div>
    </div>
  </div>

  <span id="holYear" style="display:none"></span>
  <div id="holidayList" style="display:none"></div>

</div><!-- /pg-settings -->

</div><!-- /content -->

<div class="toasts" id="toasts"></div>

<script>
// ════════════════════════════════════════════════════════
//  FİREBASE
// ════════════════════════════════════════════════════════
const FIREBASE_URL='https://grid-tracker-73ed2-default-rtdb.europe-west1.firebasedatabase.app';
async function fbRead(){
  try{
    const r=await fetch(FIREBASE_URL+'/gridtracker.json');
    const d=await r.json();
    return(d&&d.lastUpdated)?d:null;
  }catch(e){return null;}
}
async function fbWrite(path,data){
  try{
    await fetch(`${FIREBASE_URL}/${path}.json`,{method:'PUT',body:JSON.stringify(data)});
  }catch(e){}
}

// ════════════════════════════════════════════════════════
//  VERİ (Python bu bloğu günceller — Firebase yüklenemezse fallback)
// ════════════════════════════════════════════════════════
        // GRID_DATA_START
        window.__GRID_DATA__ = null;
        // GRID_DATA_END

// ════════════════════════════════════════════════════════
//  STATE
// ════════════════════════════════════════════════════════
let D = window.__GRID_DATA__ || null;
let S = {
  symMode: 'bot',
  settings: {
    commissionRate: 1, gridInterval: 0, lotTolerance: 0,
    botSymbols: [],
    costs: [
      {name:'Hesap İşletim Ücreti',amount:95},{name:'Algo',amount:1095},
      {name:'IQ Terminal',amount:980},{name:'Tek Kademe',amount:282},
      {name:'Otomatik Emir',amount:420},{name:'Endeks Veri',amount:26},
    ]
  },
  month: null,
};

// ════════════════════════════════════════════════════════
//  PERSIST
// ════════════════════════════════════════════════════════
function save(){ try{ localStorage.setItem('gt_s',JSON.stringify(S.settings)); }catch(e){} }
function load(){ try{
  const s=localStorage.getItem('gt_s'); if(s) S.settings={...S.settings,...JSON.parse(s)};
  if(!Array.isArray(S.settings.botSymbols)) S.settings.botSymbols=[];
  if(D&&D.settings&&D.settings.costs) S.settings.costs=D.settings.costs;
}catch(e){} }

// ════════════════════════════════════════════════════════
//  NAV
// ════════════════════════════════════════════════════════
function go(id,btn){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('on'));
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('on'));
  document.getElementById(id).classList.add('on');
  btn.classList.add('on');
  if(id==='pg-settings'){ renderSettings(); loadAutoSettings(); loadHolidays(); }
  if(id==='pg-monthly') renderMonthly();
  if(id==='pg-overall') renderOverall();
  if(id==='pg-trades') renderTrades();
}

// ════════════════════════════════════════════════════════
//  UTILS
// ════════════════════════════════════════════════════════
const $=id=>document.getElementById(id);
const fx=n=>(n==null||isNaN(n))?'—':Number(n).toLocaleString('tr-TR',{minimumFractionDigits:2,maximumFractionDigits:2});
const fxi=n=>(n==null||isNaN(n))?'—':Math.round(Number(n)).toLocaleString('tr-TR'); // tam sayı, kuruşsuz
const fxs=n=>n==null?'—':((n>=0?'+':'')+fx(n));
const fxsi=n=>n==null?'—':((n>=0?'+':'')+fxi(Math.abs(n))); // tam sayı signed
const setpnl=(id,v)=>{ const e=$(id); if(!e)return; e.textContent=fx(v)+' ₺'; e.className='kpi-val '+(v>0?'pos':v<0?'neg':''); };

function toast(msg,type='inf',dur=2800){
  const c=$('toasts'),d=document.createElement('div');
  d.className=`toast ${type}`; d.textContent=msg; c.appendChild(d);
  setTimeout(()=>{d.style.opacity='0';d.style.transition='opacity .3s';setTimeout(()=>d.remove(),300);},dur);
}

// ════════════════════════════════════════════════════════
//  MARKET STATUS
// ════════════════════════════════════════════════════════
function updateMarket(){
  const n=new Date(),d=n.getDay(),h=n.getHours(),m=n.getMinutes(),t=h*100+m;
  const dot=$('mDot'),txt=$('mTxt');
  if(d===0||d===6){dot.className='dot off';txt.textContent='Kapalı';return;}
  if(t>=1000&&t<=1800){dot.className='dot';txt.textContent=`${String(h).padStart(2,'0')}:${String(m).padStart(2,'0')}`;}
  else if(t>1800&&t<=1835){dot.className='dot';txt.textContent='Kapanış';}
  else{dot.className='dot off';txt.textContent='Kapalı';}
}

// ════════════════════════════════════════════════════════
//  MONTHLY
// ════════════════════════════════════════════════════════
function getMonthStats(y,mo){
  const dl=(D&&D.dailyLog)||{};
  const key=`${y}-${String(mo+1).padStart(2,'0')}`;
  const days=Object.keys(dl).filter(d=>d.startsWith(key)).sort();
  let gross=0,comm=0,net=0,tdays=0,rows=[];
  days.forEach(dk=>{
    const d=dl[dk];
    gross+=d.grossProfit||0; comm+=d.commission||0; net+=d.netProfit||0;
    if(d.trades>0) tdays++;
    rows.push(d);
  });
  const oh=(D&&D.overallHistory||[]).filter(h=>h.date.startsWith(key)).sort((a,b)=>a.date.localeCompare(b.date));
  const odiff=oh.length>=2?oh[oh.length-1].amount-oh[0].amount:0;
  return{gross,comm,net,tdays,rows,odiff};
}

// ════════════════════════════════════════════════════════
//  HOME
// ════════════════════════════════════════════════════════
function renderHome(){
  if(!D){
    const al=$('syncAlert'); al.style.display='flex'; al.className='alert err';
    $('syncMsg').textContent='Veri yok — python grid_tracker_service.py --now komutunu çalıştırın'; return;
  }

  // Sync info
  $('syncTime').textContent=D.lastUpdated?D.lastUpdated.slice(11,16):'—';
  const stale=Math.floor((Date.now()-new Date(D.lastUpdated).getTime())/60000);
  const al=$('syncAlert');
  if(stale>120){
    al.style.display='flex'; al.className='alert warn';
    $('syncMsg').textContent=`Son güncelleme ${stale} dk önce`;
  } else { al.style.display='none'; }

  // Overall — yuvarlak, kuruşsuz
  const ov=D.todayOverall||0;
  $('h-overall').innerHTML=fxi(ov)+'<span style="font-size:22px;font-weight:400;-webkit-text-fill-color:rgba(148,163,184,.55);color:rgba(148,163,184,.55)"> ₺</span>';
  // Tarih - TR formatı
  const td=D.today||'';
  $('h-overall-date').textContent=td?td.split('-').reverse().join('.'):'—';
  const hist=(D.overallHistory||[]).filter(h=>h.date<D.today).sort((a,b)=>b.date.localeCompare(a.date));
  const prev=hist[0]||null;
  const de=$('h-overall-diff');
  if(prev){
    const diff=ov-prev.amount;
    de.textContent=fxsi(diff)+' ₺';
    de.className='pf-stat-val '+(diff>0?'pos':diff<0?'neg':'');
  }else{ de.textContent='—'; de.className='pf-stat-val'; }

  // P&L
  const tp=D.todayProfit||{};
  setpnl('h-gross',tp.totalGross||0);
  $('h-comm').textContent=fx(tp.totalCommission||0)+' ₺';
  setpnl('h-net',tp.totalNet||0);

  // Trades
  const dl=(D.dailyLog||{})[D.today]||{};
  $('h-trades').textContent=dl.trades||0;
  $('h-trades-sub').textContent=`${dl.buys||0} alış · ${dl.sells||0} satış`;

  // Monthly
  const now=new Date();
  const ms=getMonthStats(now.getFullYear(),now.getMonth());
  setpnl('h-mnet',ms.net);
  $('h-mdays').textContent=ms.tdays+'. işlem günü';

  // Lisans
  renderLicense(ms.comm);

  // Sym table
  renderSymTable();
}

function renderLicense(monthComm){
  const total=S.settings.costs.reduce((a,c)=>a+(parseFloat(c.amount)||0),0);
  const target=total*3;
  const ratio=target>0?monthComm/target:0;
  const pct=Math.min(ratio*100,100);

  const b=$('licBanner'); b.style.display='block';
  $('licTarget').textContent=fx(target)+' ₺';
  $('licPct').textContent='%'+Math.round(pct);
  $('licBar').style.width=pct+'%';

  let cls,icon,title,sub;
  if(ratio>=1){ cls='lic-ok';icon='🏆';title='MatriksIQ Lisansı Ücretsiz!';sub=`🎉 MatriksIQ Terminal ve Lisans Ücretleriniz BEDAVA..!  ·  ${fx(monthComm)} ₺`; $('licBar').className='lic-glow'; }
  else if(ratio>=0.5){ cls='lic-partial';icon='⚡';title=`Hedefe ulaşıldı`;sub=`${fx(target-monthComm)} ₺ daha gerekiyor`; $('licBar').className='lic-glow'; }
  else{ cls='lic-no';icon='◎';title=`Bu ay komisyon`;sub=`${fx(monthComm)} ₺ — ${fx(target-monthComm)} ₺ eksik`; $('licBar').className='lic-glow'; }

  b.className=`lic-banner ${cls}`;
  $('licIcon').textContent=icon; $('licTitle').textContent=title; $('licSub').textContent=sub;
}

function setSymMode(mode,el){
  S.symMode=mode;
  document.querySelectorAll('.sym-tab').forEach(b=>b.classList.remove('on'));
  if(el) el.classList.add('on');
  renderSymTable();
}
function renderSymTable(){
  const syms=D?D.todayProfit?D.todayProfit.bySymbol||{}:{}:{};
  let keys=Object.keys(syms).sort();
  const isBot=S.symMode==='bot';
  const botList=(S.settings.botSymbols||[]).map(s=>s.toUpperCase());
  if(isBot && botList.length) keys=keys.filter(k=>botList.includes(k.toUpperCase()));
  const total=keys.length;
  $('h-sym-count').textContent=`${total} hisse`;
  const tb=$('h-sym-tbl');
  if(!keys.length){
    const msg=isBot&&botList.length===0?'GridBot hissesi tanımlanmamış':'Bugün işlem yok';
    tb.innerHTML=`<tr><td colspan="6"><div class="empty"><div class="empty-icon">—</div><div class="empty-text">${msg}</div></div></td></tr>`; return;
  }
  tb.innerHTML=keys.map(k=>{
    const s=syms[k];
    const nc=s.netProfit>0?'td-pos':s.netProfit<0?'td-neg':'';
    return `<tr>
      <td class="td-sym">${k}</td>
      <td class="td-dim">${s.buyCount}</td>
      <td class="td-sell">${s.sellCount}</td>
      <td>${fx(s.avgBuy)}</td>
      <td>${fx(s.avgSell)}</td>
      <td class="${nc}">${fxs(s.netProfit)} ₺</td>
    </tr>`;
  }).join('');
}

// ════════════════════════════════════════════════════════
//  TRADES
// ════════════════════════════════════════════════════════
function renderTrades(){
  if(!D) return;
  const fSym=$('fSym').value, fType=$('fType').value;
  const syms=[...new Set((D.allTrades||[]).map(t=>t.symbol))].sort();
  const cur=$('fSym').value;
  $('fSym').innerHTML='<option value="">Tüm Hisseler</option>'+syms.map(s=>`<option ${s===cur?'selected':''}>${s}</option>`).join('');
  let trs=[...(D.allTrades||[])];
  if(fSym) trs=trs.filter(t=>t.symbol===fSym);
  if(fType) trs=trs.filter(t=>t.type===fType);
  trs.sort((a,b)=>b.datetime.localeCompare(a.datetime));
  const tb=$('trades-tbl');
  if(!trs.length){ tb.innerHTML=`<tr><td colspan="7"><div class="empty"><div class="empty-icon">↕</div><div class="empty-text">İşlem yok</div></div></td></tr>`; return; }
  tb.innerHTML=trs.map(t=>`<tr>
    <td class="td-time">${t.time}</td>
    <td class="td-sym">${t.symbol}</td>
    <td><span class="pill ${t.type==='Alış'?'pill-b':'pill-s'}">${t.type==='Alış'?'AL':'SAT'}</span></td>
    <td class="td-lot">${Number(t.qty).toLocaleString('tr-TR')}</td>
    <td class="td-price">${fx(t.execPrice)}</td>
    <td class="td-amount">${fx(t.execAmount)}</td>
    <td class="td-amb">${fx(t.commission)}</td>
  </tr>`).join('');
}

function exportCSV(){
  if(!D) return;
  const h=['Tarih','Saat','Hisse','İşlem','Lot','Fiyat','Tutar','Komisyon'];
  const rows=(D.allTrades||[]).map(t=>[t.date,t.time,t.symbol,t.type,t.qty,t.execPrice,t.execAmount,t.commission].join(';'));
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob(['\\ufeff'+[h.join(';'),...rows].join('\\n')],{type:'text/csv'}));
  a.download='grid-islemler.csv'; a.click();
  toast('CSV indirildi','ok');
}

// ════════════════════════════════════════════════════════
//  MONTHLY
// ════════════════════════════════════════════════════════
function renderMonthly(){
  const now=new Date();
  if(!S.month) S.month={y:now.getFullYear(),m:now.getMonth()};
  const MONTHS=['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık'];
  $('monthLbl').innerHTML=`<span class="month-label-m">${MONTHS[S.month.m]}</span><span class="month-label-y">${S.month.y}</span>`;
  const st=getMonthStats(S.month.y,S.month.m);
  $('m-days').textContent=st.tdays+'. işlem günü';
  setpnl('m-gross',st.gross); $('m-comm').textContent=fx(st.comm)+' ₺'; setpnl('m-net',st.net);
  const avg=st.tdays?st.net/st.tdays:0;
  const ea=$('m-avg'); if(ea){ea.textContent=fx(avg)+' ₺';ea.className='kpi-val kpi-muted';}
  const dl2=(D&&D.dailyLog)||{};const mmap={};Object.keys(dl2).forEach(dt=>{const mk=dt.slice(0,7);if(!mmap[mk])mmap[mk]=0;mmap[mk]+=(dl2[dt].netProfit||0);});const mvals=Object.values(mmap);const mavg=mvals.length?mvals.reduce((a,b)=>a+b,0)/mvals.length:0;
  const em=$('m-mavg'); if(em){em.textContent=fx(mavg)+' ₺';em.className='kpi-val kpi-muted';}
  const od=$('m-odiff'); od.textContent=(st.odiff>=0?'+':'')+fxi(st.odiff)+' ₺'; od.className='kpi-val '+(st.odiff>0?'pos':st.odiff<0?'neg':'');
  const tb=$('m-tbl');
  if(!st.rows.length){ tb.innerHTML=`<tr><td colspan="6"><div class="empty"><div class="empty-icon">◫</div><div class="empty-text">Bu ay veri yok</div></div></td></tr>`; return; }
  tb.innerHTML=st.rows.map(d=>{
    const nc=d.netProfit>0?'td-pos':d.netProfit<0?'td-neg':'';
    return `<tr>
      <td style="color:#7b92b5;font-weight:600;font-size:11px">${d.date.slice(5).replace('-','.')}</td>
      <td class="td-dim">${d.trades}</td>
      <td class="${d.grossProfit>=0?'td-pos':'td-neg'}">${fxs(d.grossProfit)}</td>
      <td class="td-amb">${fx(d.commission)}</td>
      <td class="${nc}" style="font-weight:600">${fxs(d.netProfit)}</td>
      <td style="color:#4e6080;font-size:10px">${d.overall?fx(d.overall)+' ₺':'—'}</td>
    </tr>`;
  }).join('');
}
function changeMonth(dir){
  S.month.m+=dir;
  if(S.month.m<0){S.month.m=11;S.month.y--;}
  if(S.month.m>11){S.month.m=0;S.month.y++;}
  renderMonthly();
}

// ════════════════════════════════════════════════════════
//  OVERALL
// ════════════════════════════════════════════════════════
function renderOverall(){
  const oh=D?[...(D.overallHistory||[])]:[];
  oh.sort((a,b)=>b.date.localeCompare(a.date));
  const tb=$('o-tbl');
  if(!oh.length){ tb.innerHTML=`<tr><td colspan="5"><div class="empty"><div class="empty-icon">◉</div><div class="empty-text">Kayıt yok — Python servisi her gün otomatik ekler</div></div></td></tr>`; return; }
  tb.innerHTML=oh.map((h,i)=>{
    const prev=oh[i+1];
    const diff=prev?h.amount-prev.amount:null;
    const dc=diff!==null?(diff>=0?'td-pos':'td-neg'):'';
    const diffTxt=diff!==null?(diff>=0?'+':'')+fxi(diff)+' ₺':'—';
    return `<tr>
      <td style="color:#7b92b5;font-weight:600;font-size:11px">${h.date.split('-').reverse().join('.')}</td>
      <td style="color:#c4cfe0;font-weight:600;font-family:var(--mono)">${fxi(h.amount)} ₺</td>
      <td class="${dc}" style="font-weight:${diff!==null?'600':'400'}">${diffTxt}</td>
      <td style="color:#4e6080;font-family:var(--sans);font-size:10px">${h.note||'—'}</td>
      <td><button class="btn btn-danger btn-xs" onclick="delOverall('${h.date}')" style="padding:2px 6px;font-size:10px;opacity:.6">✕</button></td>
    </tr>`;
  }).join('');
}
function toggleOForm(){ const f=$('oForm'); f.style.display=f.style.display==='none'?'block':'none'; if(f.style.display==='block') $('oDate').value=new Date().toISOString().split('T')[0]; }
function saveOverall(){
  if(!D) return;
  const date=$('oDate').value, amt=parseFloat($('oAmt').value), note=$('oNote').value||'Manuel';
  if(!date||!amt){ toast('Tarih ve tutar gerekli','err'); return; }
  if(!D.overallHistory) D.overallHistory=[];
  const idx=D.overallHistory.findIndex(h=>h.date===date);
  if(idx>=0) D.overallHistory[idx]={date,amount:amt,note};
  else D.overallHistory.push({date,amount:amt,note});
  D.overallHistory.sort((a,b)=>a.date.localeCompare(b.date));
  $('oForm').style.display='none'; renderOverall(); toast('Kaydedildi','ok');
}
function delOverall(date){ if(!D) return; D.overallHistory=D.overallHistory.filter(h=>h.date!==date); renderOverall(); }

// ════════════════════════════════════════════════════════
//  BOT HİSSELERİ
// ════════════════════════════════════════════════════════
function addBotSym(){
  const sym=$('botSymInput').value.trim().toUpperCase();
  if(!sym){ toast('Hisse kodu girin','err'); return; }
  if((S.settings.botSymbols||[]).includes(sym)){ toast(`${sym} zaten listede`,'inf'); return; }
  if(!S.settings.botSymbols) S.settings.botSymbols=[];
  S.settings.botSymbols.push(sym);
  save(); renderBotSymList(); $('botSymInput').value='';
  toast(`${sym} bot listesine eklendi`,'ok');
}
function removeBotSym(i){ S.settings.botSymbols.splice(i,1); save(); renderBotSymList(); renderSymTable(); }
function renderBotSymList(){
  const list=S.settings.botSymbols||[];
  const el=$('botSymList');
  if(!list.length){
    el.innerHTML=`<div style="font-size:10px;color:#3d5070;padding:4px 0;font-style:italic">Henüz hisse eklenmedi</div>`; return;
  }
  el.innerHTML=`<div style="display:flex;flex-wrap:wrap;gap:5px;padding-top:4px">`+
    list.map((s,i)=>`<div style="display:flex;align-items:center;gap:4px;background:rgba(59,130,246,.07);border:1px solid rgba(59,130,246,.16);border-radius:5px;padding:3px 8px">
      <span style="font-family:var(--mono);font-size:11px;font-weight:600;color:#7aa6e0">${s}</span>
      <button onclick="removeBotSym(${i})" style="background:none;border:none;cursor:pointer;color:#3d5070;font-size:10px;padding:0;line-height:1" title="Kaldır">✕</button>
    </div>`).join('')+`</div>`;
}

// ════════════════════════════════════════════════════════
//  SETTINGS
// ════════════════════════════════════════════════════════
function renderSettings(){
  $('sComm').value=S.settings.commissionRate;
  $('sGrid').value=S.settings.gridInterval;
  $('sTol').value=S.settings.lotTolerance;
  renderCostList(); renderCostSummary(); renderBotSymList();
}
function saveSettings(){
  S.settings.commissionRate=parseFloat($('sComm').value)||1;
  S.settings.gridInterval=parseFloat($('sGrid').value)||0;
  S.settings.lotTolerance=parseFloat($('sTol').value)||0;
  save(); toast('Kaydedildi','ok'); renderHome();
}
function renderCostList(){
  $('costList').innerHTML=S.settings.costs.map((c,i)=>`
    <div class="cost-row">
      <div class="cost-name">${c.label||c.name||'—'}</div>
      <input class="cost-edit" type="number" value="${c.amount}"
        onchange="S.settings.costs[${i}].amount=parseFloat(this.value)||0;renderCostSummary();">
      <span class="cost-unit">₺</span>
    </div>`).join('');
}
function addCost(){ S.settings.costs.push({name:'Yeni Kalem',amount:0}); renderCostList(); renderCostSummary(); }
function saveCosts(){
  // Kalem adlarını güncelle (input value yerine mevcut değerleri kullan)
  save();
  if(D) D.settings=S.settings;
  renderCostSummary(); renderLicense(getMonthStats(new Date().getFullYear(),new Date().getMonth()).comm);
  toast('Kaydedildi','ok');
}
function renderCostSummary(){
  const total=S.settings.costs.reduce((a,c)=>a+(parseFloat(c.amount)||0),0);
  const target=total*3;
  const ms=getMonthStats(new Date().getFullYear(),new Date().getMonth());
  $('sTotalCost').textContent=fx(total)+' ₺';
  $('sTarget').textContent=fx(target)+' ₺';
  $('sEarned').textContent=fx(ms.comm)+' ₺';
}

// ════════════════════════════════════════════════════════
//  SABAH OTOMASYONU AYARLARI
// ════════════════════════════════════════════════════════
const AUTO_API = 'http://localhost:5050';
const IS_LOCAL = location.hostname === 'localhost' || location.hostname === '127.0.0.1';

async function loadAutoSettings(){
  if(!IS_LOCAL){ $('autoStatus').className='alert warn'; $('autoStatus').textContent='⚠ Otomasyon ayarları sadece yerel ağdan erişilebilir'; $('autoStatus').style.display='flex'; return; }
  try{
    const r = await fetch(AUTO_API+'/api/morning-settings');
    if(!r.ok) throw new Error();
    const d = await r.json();
    $('autoTime').value = d.time || '09:15';
    $('autoWake').textContent = d.wake_time || '—';
    $('autoStatus').style.display='none';
  }catch{
    $('autoStatus').className='alert warn';
    $('autoStatus').textContent='⚠ Ayarlar sunucusu çalışmıyor (automation_server.py)';
    $('autoStatus').style.display='flex';
  }
}

async function saveAutoSettings(){
  const t = $('autoTime').value;
  if(!t){ toast('Saat seçin','err'); return; }
  try{
    const r = await fetch(AUTO_API+'/api/morning-settings',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({time: t})
    });
    const d = await r.json();
    if(!r.ok){ toast(d.error||'Hata','err'); return; }
    $('autoWake').textContent = d.wake_time;
    $('biosWakeDisplay').textContent = d.wake_time;
    $('biosReminder').style.display = 'block';
    toast(d.task_ok ? 'Görev güncellendi ✓' : 'Kaydedildi (görev hatası)', d.task_ok?'ok':'warn');
  }catch{
    toast('Sunucuya ulaşılamadı','err');
  }
}

async function loadHolidays(){
  if(!IS_LOCAL){ $('holidayList').innerHTML='<div style="font-size:11px;color:var(--text3)">Tatil listesi sadece yerel ağdan görüntülenebilir.</div>'; return; }
  const year = new Date().getFullYear();
  $('holYear').textContent = year;
  try{
    const r = await fetch(AUTO_API+'/api/holidays/'+year);
    if(!r.ok) throw new Error();
    const items = await r.json();
    const today = new Date().toISOString().split('T')[0];
    const upcoming = items.filter(x => x.date >= today);
    if(!upcoming.length){ $('holidayList').innerHTML='<div style="color:var(--text3);font-size:12px">Bu yıl için tatil bulunamadı.</div>'; return; }
    $('holidayList').innerHTML = upcoming.map(x=>{
      const parts=x.date.split('-');
      const label=parts[2]+'.'+parts[1]+'.'+parts[0];
      return '<div style="display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px solid var(--border)">'+
        '<span style="font-size:12px;color:var(--text)">'+x.name+'</span>'+
        '<span style="font-family:var(--mono);font-size:11px;color:var(--text3)">'+label+'</span>'+
      '</div>';
    }).join('');
  }catch{
    $('holidayList').innerHTML='<div style="color:var(--text3);font-size:12px">Sunucu çalışmıyor — tatiller gösterilemiyor.</div>';
  }
}

// ════════════════════════════════════════════════════════
//  DATA OPS
// ════════════════════════════════════════════════════════
function exportJSON(){
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([JSON.stringify({gridData:D,settings:S.settings,manual:S.manual},null,2)],{type:'application/json'}));
  a.download=`grid-backup-${new Date().toISOString().split('T')[0]}.json`; a.click();
  toast('Yedek indirildi','ok');
}
function importJSON(e){
  const f=e.target.files[0]; if(!f) return;
  const r=new FileReader();
  r.onload=ev=>{
    try{
      const d=JSON.parse(ev.target.result);
      if(d.settings) S.settings={...S.settings,...d.settings};
      if(d.manual) S.manual=d.manual;
      save(); renderHome(); renderSettings(); toast('Yedek yüklendi','ok');
    }catch{ toast('Yükleme hatası','err'); }
  };
  r.readAsText(f);
}
function clearAll(){
  if(!confirm('Tüm yerel veriler (ayarlar, bot hisseleri) silinecek. Emin misiniz?')) return;
  S.settings.botSymbols=[]; localStorage.removeItem('gt_s');
  renderBotSymList(); renderSymTable(); toast('Sıfırlandı','inf');
}

// ════════════════════════════════════════════════════════
//  INIT
// ════════════════════════════════════════════════════════
document.addEventListener('DOMContentLoaded',async()=>{
  load();
  updateMarket(); setInterval(updateMarket,30000);
  $('oDate').value=new Date().toISOString().split('T')[0];

  // Firebase'den güncel veriyi yükle
  const fb=await fbRead();
  if(fb){
    D=fb;
    if(D.settings&&D.settings.costs) S.settings.costs=D.settings.costs;
  }
  renderHome();

  // Her 5 dakikada bir Firebase'i kontrol et
  setInterval(async()=>{
    const fb=await fbRead();
    if(fb&&(!D||fb.lastUpdated>D.lastUpdated)){
      D=fb;
      renderHome();
      toast('Veri güncellendi','ok',2000);
    }
  },300000);

  // ════ AUTO_DATA_INJECT ════
});
</script>
</body>
</html>
'''

# ──────────────────────────────────────────────────────────
#  YAPILANDIRMA
# ──────────────────────────────────────────────────────────
DESKTOP    = Path.home() / 'Desktop'
FILE1      = DESKTOP / '1.xlsx'
FILE2      = DESKTOP / '2.xlsx'

SCRIPT_DIR = Path(__file__).parent
HTML_FILE  = SCRIPT_DIR / 'bist_tracker.html'
DATA_JSON  = SCRIPT_DIR / 'data.json'
LOG_FILE   = SCRIPT_DIR / 'grid_tracker.log'

NORMAL_CLOSE    = (18, 0)
HALF_DAY_CLOSE  = (13, 0)   # Arife günleri BIST 13:00'da kapanır
OFFSET_MIN      = 35
COMMISSION_RATE = 1 / 10000   # Her alış VE satış için ayrı ayrı
FIREBASE_URL    = 'https://grid-tracker-73ed2-default-rtdb.europe-west1.firebasedatabase.app'

# Arife günleri: BIST 13:00'da kapanır → akşam otomasyonu 13:35'te çalışır
# Sabit ulusal bayramların arife günü: sadece Cumhuriyet Bayramı (29 Eki) öncesi 28 Eki
ARIFE_DAYS = {
    # 2025
    '2025-03-28',  # Ramazan Bayramı arefe
    '2025-06-05',  # Kurban Bayramı arefe
    '2025-10-28',  # Cumhuriyet Bayramı arefe
    # 2026
    '2026-03-19',  # Ramazan Bayramı arefe
    '2026-05-26',  # Kurban Bayramı arefe
    '2026-10-28',  # Cumhuriyet Bayramı arefe
    # 2027
    '2027-03-08',  # Ramazan Bayramı arefe
    '2027-05-15',  # Kurban Bayramı arefe
    '2027-10-28',  # Cumhuriyet Bayramı arefe
    # 2028
    '2028-02-25',  # Ramazan Bayramı arefe
    '2028-05-03',  # Kurban Bayramı arefe
    '2028-10-28',  # Cumhuriyet Bayramı arefe
}

# Resmi tatil günleri: BIST kapalı
HOLIDAYS = {
    # ── 2025 ──────────────────────────────────────────────
    '2025-01-01',                                           # Yılbaşı
    '2025-03-29','2025-03-30','2025-03-31',                 # Ramazan Bayramı
    '2025-04-23',                                           # Ulusal Egemenlik ve Çocuk Bayramı
    '2025-05-01',                                           # Emek ve Dayanışma Günü
    '2025-05-19',                                           # Gençlik ve Spor Bayramı
    '2025-06-06','2025-06-07','2025-06-08','2025-06-09',    # Kurban Bayramı
    '2025-07-15',                                           # Demokrasi ve Milli Birlik Günü
    '2025-08-30',                                           # Zafer Bayramı
    '2025-10-29',                                           # Cumhuriyet Bayramı
    # ── 2026 ──────────────────────────────────────────────
    '2026-01-01',                                           # Yılbaşı
    '2026-03-20','2026-03-21','2026-03-22',                 # Ramazan Bayramı
    '2026-04-23',                                           # Ulusal Egemenlik ve Çocuk Bayramı
    '2026-05-01',                                           # Emek ve Dayanışma Günü
    '2026-05-19',                                           # Gençlik ve Spor Bayramı
    '2026-05-27','2026-05-28','2026-05-29','2026-05-30',    # Kurban Bayramı
    '2026-07-15',                                           # Demokrasi ve Milli Birlik Günü
    '2026-08-30',                                           # Zafer Bayramı
    '2026-10-29',                                           # Cumhuriyet Bayramı
    # ── 2027 ──────────────────────────────────────────────
    '2027-01-01',                                           # Yılbaşı
    '2027-03-09','2027-03-10','2027-03-11',                 # Ramazan Bayramı
    '2027-04-23',                                           # Ulusal Egemenlik ve Çocuk Bayramı
    '2027-05-01',                                           # Emek ve Dayanışma Günü
    '2027-05-16','2027-05-17','2027-05-18','2027-05-19',    # Kurban Bayramı (19 May = Gençlik Bayramı ile çakışıyor)
    '2027-07-15',                                           # Demokrasi ve Milli Birlik Günü
    '2027-08-30',                                           # Zafer Bayramı
    '2027-10-29',                                           # Cumhuriyet Bayramı
    # ── 2028 ──────────────────────────────────────────────
    '2028-01-01',                                           # Yılbaşı
    '2028-02-26','2028-02-27','2028-02-28',                 # Ramazan Bayramı
    '2028-04-23',                                           # Ulusal Egemenlik ve Çocuk Bayramı
    '2028-05-01',                                           # Emek ve Dayanışma Günü
    '2028-05-04','2028-05-05','2028-05-06','2028-05-07',    # Kurban Bayramı
    '2028-05-19',                                           # Gençlik ve Spor Bayramı
    '2028-07-15',                                           # Demokrasi ve Milli Birlik Günü
    '2028-08-30',                                           # Zafer Bayramı
    '2028-10-29',                                           # Cumhuriyet Bayramı
}

DEFAULT_COSTS = [
    {'name':'Hesap İşletim Ücreti','amount':95},
    {'name':'Algo',                'amount':1095},
    {'name':'IQ Terminal',         'amount':980},
    {'name':'Tek Kademe',          'amount':282},
    {'name':'Otomatik Emir',       'amount':420},
    {'name':'Endeks Veri',         'amount':26},
]

# ──────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger('GridTracker')

# ──────────────────────────────────────────────────────────
#  TAKVİM
# ──────────────────────────────────────────────────────────
def is_trading_day(d=None):
    if d is None: d = date.today()
    if d.weekday() >= 5: return False
    if d.strftime('%Y-%m-%d') in HOLIDAYS: return False
    return True

def is_arife(d=None):
    if d is None: d = date.today()
    return d.strftime('%Y-%m-%d') in ARIFE_DAYS

def is_last_bist_day_of_month(d=None):
    """Verilen gün, o ayın son BIST işlem günü mü?"""
    if d is None: d = date.today()
    check = d + timedelta(days=1)
    while check.month == d.month:
        if is_trading_day(check):
            return False
        check += timedelta(days=1)
    return True

def calc_monthly_kar(excel_date, overall, oh, birikim_tx=None):
    """
    Aylık net karı hesaplar:
    Bu ayın son günü overall'ı − Önceki ayın son günü overall'ı − O ay net sermaye hareketi
    2026-03'ten itibaren sermaye hareketleri (birikimTx) hesaba katılır.
    """
    month_key = excel_date[:7]                          # örn: '2026-03'
    y, m = int(month_key[:4]), int(month_key[5:7])
    if m == 1:
        prev_key = f'{y-1}-12'
    else:
        prev_key = f'{y}-{m-1:02d}'
    prev_entries = sorted(
        [h for h in oh if h['date'].startswith(prev_key)],
        key=lambda x: x['date']
    )
    if not prev_entries:
        return None   # Önceki ay verisi yok, hesaplanamaz
    prev_last = prev_entries[-1]['amount']
    raw_change = overall - prev_last

    # 2026-03'ten itibaren sermaye hareketlerini çıkar
    net_capital = 0
    if birikim_tx and month_key >= '2026-03':
        for tx in birikim_tx:
            if tx.get('exclude'):
                continue
            tx_date = tx.get('date', '')
            if not tx_date.startswith(month_key):
                continue
            tip = tx.get('tip', '')
            miktar = tx.get('miktar', 0)
            if tip == 'giriş':
                net_capital += miktar
            elif tip == 'çıkış':
                net_capital -= miktar
        if net_capital:
            log.info(f'Sermaye düzeltmesi ({month_key}): net={net_capital:+,} ₺')

    return round(raw_change - net_capital)

def get_run_time(d=None):
    if d is None: d = date.today()
    h, m = HALF_DAY_CLOSE if is_arife(d) else NORMAL_CLOSE
    total = h * 60 + m + OFFSET_MIN
    return (total // 60, total % 60)

def seconds_until_run():
    now   = datetime.now()
    today = now.date()
    if is_trading_day(today):
        rh, rm = get_run_time(today)
        run_t  = now.replace(hour=rh, minute=rm, second=0, microsecond=0)
        if now < run_t:
            d = (run_t - now).total_seconds()
            log.info(f"Bugün çalışma: {rh:02d}:{rm:02d} — {d/60:.1f} dk sonra")
            return d
    check = today + timedelta(days=1)
    for _ in range(10):
        if is_trading_day(check):
            rh, rm = get_run_time(check)
            nxt = datetime.combine(check, datetime.min.time()).replace(
                hour=rh, minute=rm, second=0, microsecond=0)
            d = (nxt - now).total_seconds()
            log.info(f"Sonraki çalışma: {check} {rh:02d}:{rm:02d} — {d/3600:.1f} saat sonra")
            return d
        check += timedelta(days=1)
    return 86400

# ──────────────────────────────────────────────────────────
#  EXCEL OKUMA
# ──────────────────────────────────────────────────────────
def sf(v, t=float, d=0.0):
    try: return t(v) if v is not None else d
    except: return d

def ss(v, d=''):
    try: return str(v).strip() if v is not None else d
    except: return d

def read_file1(path):
    if not path.exists():
        log.error(f"Dosya yok: {path}"); return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    trades = []
    for i, r in enumerate(rows):
        if i == 0 or not r[5]: continue
        raw_dt, raw_t = r[21], r[19]
        if isinstance(raw_dt, datetime):
            date_s = raw_dt.strftime('%Y-%m-%d')
            dt_s   = raw_dt.strftime('%Y-%m-%d %H:%M:%S')
        else:
            s = ss(raw_dt); date_s = s[:10]; dt_s = s[:19]
        time_s = raw_t.strftime('%H:%M:%S') if isinstance(raw_t, datetime) else ss(raw_t)[:8]

        qty        = sf(r[8])
        price      = sf(r[9])
        exec_qty   = sf(r[12]) or qty
        amount     = sf(r[14])
        exec_price = sf(r[16]) or price
        exec_amt   = sf(r[18]) or amount
        comm       = round(exec_amt * COMMISSION_RATE, 6)

        trades.append({
            'symbol':     ss(r[5]).upper(),
            'type':       ss(r[7]),          # 'Alış' veya 'Satış'
            'qty':        int(qty),
            'price':      price,
            'status':     ss(r[10]),
            'execQty':    int(exec_qty),
            'amount':     amount,
            'execPrice':  exec_price,
            'execAmount': exec_amt,
            'commission': comm,
            'time':       time_s,
            'date':       date_s,
            'datetime':   dt_s,
            'referans':   ss(r[6]),
        })
    log.info(f"Dosya 1: {len(trades)} işlem okundu")
    return trades

def read_file2(path):
    if not path.exists():
        log.error(f"Dosya yok: {path}"); return 0.0
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        kod  = ss(row[0]).upper()
        desc = ss(row[1]).lower()
        val  = row[2]
        if kod == 'OAL' or ('t2' in desc and 'overall' in desc):
            v = sf(val)
            if v: log.info(f"T2 Overall: {v:,.2f} ₺"); return v
    for row in ws.iter_rows(values_only=True):
        if ss(row[0]).upper() == 'OAC':
            return sf(row[2])
    log.warning("T2 Overall bulunamadı"); return 0.0

# ──────────────────────────────────────────────────────────
#  KAR HESAPLAMA
# ──────────────────────────────────────────────────────────
def calc_profit(trades, carried_over=None):
    """
    FIFO eşleştirme.
    carried_over : {sembol: [önceki günden devredilen eşleşmemiş alış işlemleri]}
    Kural: devredilen alışın komisyonu dünkü hesapta zaten düşüldü,
           bugün tekrar sayılmaz; yalnızca bugünkü satış komisyonu sayılır.
    """
    if carried_over is None:
        carried_over = {}

    all_syms  = list(dict.fromkeys(
        list(carried_over.keys()) + [t['symbol'] for t in trades]
    ))
    by_symbol = {}

    for sym in all_syms:
        sym_trades = sorted(
            [t for t in trades if t['symbol'] == sym],
            key=lambda x: x['datetime']
        )
        today_buys  = [t for t in sym_trades if t['type'] == 'Alış']
        today_sells = [t for t in sym_trades if t['type'] == 'Satış']

        # FIFO kuyruğu: önceki günden devredenler önce, sonra bugünkü alışlar
        carried   = [dict(t, carryover=True)  for t in carried_over.get(sym, [])]
        buy_queue = carried + list(today_buys)

        pairs = []
        for sell in today_sells:
            sell_remaining = sell['execQty']
            sell_cpl = sell['commission'] / sell['execQty']  # komisyon / lot

            while sell_remaining > 0 and buy_queue:
                buy       = buy_queue[0]
                match_qty = min(buy['execQty'], sell_remaining)
                gross     = (sell['execPrice'] - buy['execPrice']) * match_qty

                # Orantılı komisyon: satış tarafı her zaman, alış tarafı devredilmemişse
                buy_cpl  = buy['commission'] / buy['execQty']
                pair_comm = sell_cpl * match_qty
                if not buy.get('carryover'):
                    pair_comm += buy_cpl * match_qty

                pairs.append({
                    'buyPrice':  round(buy['execPrice'],  4),
                    'sellPrice': round(sell['execPrice'], 4),
                    'qty':       match_qty,
                    'gross':     round(gross,      4),
                    'comm':      round(pair_comm,  4),
                    'net':       round(gross - pair_comm, 4),
                    'buyTime':   buy['time'],
                    'sellTime':  sell['time'],
                    'buyDate':   buy.get('date', ''),
                    'carryover': bool(buy.get('carryover')),
                })

                sell_remaining -= match_qty

                if match_qty >= buy['execQty']:
                    buy_queue.pop(0)          # alış tamamen tükendi
                else:
                    # Alış kısmen tükendi → kalan miktarı güncelle
                    buy_queue[0] = dict(buy,
                        execQty=buy['execQty'] - match_qty,
                        commission=buy_cpl * (buy['execQty'] - match_qty)
                    )

        # Bugünkü komisyon = sadece bugün gerçekleşen işlemler (devredilen alış hariç)
        today_comm  = sum(t['commission'] for t in sym_trades)
        gross_total = sum(p['gross'] for p in pairs)
        net_total   = gross_total - today_comm

        avg_buy  = (sum(t['execPrice'] for t in today_buys)  / len(today_buys))  if today_buys  else 0
        avg_sell = (sum(t['execPrice'] for t in today_sells) / len(today_sells)) if today_sells else 0

        # Kalan eşleşmemiş alışlar → ertesi güne devredilecek (carryover bayrağı sıfırlanır)
        next_day_open = [dict(b, carryover=False) for b in buy_queue]

        by_symbol[sym] = {
            'symbol':        sym,
            'buyCount':      len(today_buys),
            'sellCount':     len(today_sells),
            'pairCount':     len(pairs),
            'openBuys':      sum(b['execQty'] for b in buy_queue if not b.get('carryover')),
            'carriedIn':     sum(b['execQty'] for b in carried),   # bugüne devreden lot
            'avgBuy':        round(avg_buy,  4),
            'avgSell':       round(avg_sell, 4),
            'grossProfit':   round(gross_total, 4),
            'commission':    round(today_comm,  4),
            'netProfit':     round(net_total,   4),
            'pairs':         pairs,
            'openPositions': next_day_open,         # yarına devredilecek
        }

    return {
        'bySymbol':        by_symbol,
        'totalGross':      round(sum(v['grossProfit'] for v in by_symbol.values()), 4),
        'totalCommission': round(sum(v['commission']  for v in by_symbol.values()), 4),
        'totalNet':        round(sum(v['netProfit']   for v in by_symbol.values()), 4),
        # Tüm sembollerin açık pozisyonları (yarına devir)
        'openPositions':   {sym: v['openPositions']
                            for sym, v in by_symbol.items() if v['openPositions']},
    }

# ──────────────────────────────────────────────────────────
#  FİREBASE
# ──────────────────────────────────────────────────────────
def firebase_write(payload):
    """Veriyi Firebase Realtime Database'e yazar."""
    try:
        url  = f'{FIREBASE_URL}/gridtracker.json'
        resp = requests.put(url, json=payload, timeout=15)
        if resp.status_code == 200:
            log.info('Firebase: veri yazıldı ✓')
            return True
        log.warning(f'Firebase yazma hatası: HTTP {resp.status_code}')
    except Exception as e:
        log.warning(f'Firebase bağlantı hatası: {e}')
    return False

def firebase_read():
    """Firebase'den mevcut veriyi okur."""
    try:
        url  = f'{FIREBASE_URL}/gridtracker.json'
        resp = requests.get(url, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if data:
                log.info('Firebase: mevcut veri okundu ✓')
                return data
    except Exception as e:
        log.warning(f'Firebase okuma hatası: {e}')
    return None

# ──────────────────────────────────────────────────────────
#  MEVCUT VERİYİ OKU (Firebase → HTML → data.json)
# ──────────────────────────────────────────────────────────
def load_existing():
    # 1. Firebase (güncel kaynak)
    fb = firebase_read()
    if fb:
        return fb
    # 2. HTML içindeki gömülü veri
    if HTML_FILE.exists():
        html = HTML_FILE.read_text(encoding='utf-8')
        m = re.search(
            r'//\s*GRID_DATA_START\s*\n\s*window\.__GRID_DATA__\s*=\s*(\{.*?\});\s*\n\s*//\s*GRID_DATA_END',
            html, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(1))
            except Exception as e:
                log.warning(f"HTML veri parse hatası: {e}")
    # 3. Yedek: data.json
    if DATA_JSON.exists():
        try:
            return json.loads(DATA_JSON.read_text(encoding='utf-8'))
        except:
            pass
    return {}


# ──────────────────────────────────────────────────────────
#  HTML DOSYASINI OLUŞTUR / GÜNCELLE
# ──────────────────────────────────────────────────────────
def ensure_html():
    """
    HTML_TEMPLATE içeriğinden bist_tracker.html oluşturur.
    Dosya yoksa sıfırdan yazar.
    Dosya varsa şablon değişmişse günceller, mevcut veri bloğunu korur.
    """
    import hashlib
    new_hash = hashlib.md5(HTML_TEMPLATE.encode('utf-8')).hexdigest()[:12]
    ver_tag  = f'GT_TMPL_{new_hash}'

    def build_html(data_block=None):
        """Şablondan HTML oluşturur, varsa veri bloğunu geri koyar."""
        out = HTML_TEMPLATE
        if data_block:
            out = re.sub(
                r'// GRID_DATA_START.*?// GRID_DATA_END',
                data_block, out, flags=re.DOTALL
            )
        out = out.replace(
            '// ════ AUTO_DATA_INJECT ════',
            f'// ════ AUTO_DATA_INJECT ════ {ver_tag}'
        )
        return out

    if not HTML_FILE.exists():
        HTML_FILE.write_text(build_html(), encoding='utf-8')
        log.info(f"HTML oluşturuldu: {HTML_FILE.name}")
        return

    existing = HTML_FILE.read_text(encoding='utf-8')

    # Şablon değişmemişse dokunma
    if ver_tag in existing:
        return

    # Mevcut veri bloğunu kaydet
    m = re.search(r'(// GRID_DATA_START.*?// GRID_DATA_END)', existing, re.DOTALL)
    data_block = m.group(1) if m else None

    HTML_FILE.write_text(build_html(data_block), encoding='utf-8')
    log.info(f"HTML güncellendi (sürüm: {new_hash[:8]})")

# ──────────────────────────────────────────────────────────
#  HTML'E VERİ GÖM
# ──────────────────────────────────────────────────────────
def inject_into_html(payload):
    if not HTML_FILE.exists():
        log.error(f"HTML bulunamadı: {HTML_FILE}"); return False
    html     = HTML_FILE.read_text(encoding='utf-8')
    json_str = json.dumps(payload, ensure_ascii=False, indent=2)
    new_block = (
        "// GRID_DATA_START\n"
        f"        window.__GRID_DATA__ = {json_str};\n"
        "        // GRID_DATA_END"
    )
    pattern = r'//\s*GRID_DATA_START.*?//\s*GRID_DATA_END'
    if re.search(pattern, html, re.DOTALL):
        new_html = re.sub(pattern, new_block, html, flags=re.DOTALL)
    else:
        marker   = '// ════ AUTO_DATA_INJECT ════'
        new_html = html.replace(marker, new_block + '\n        ' + marker)
    HTML_FILE.write_text(new_html, encoding='utf-8')
    log.info(f"HTML güncellendi → {HTML_FILE.name}")
    return True

# ──────────────────────────────────────────────────────────
#  ANA FONKSİYON
# ──────────────────────────────────────────────────────────
def run_once(dry_run=False):
    ensure_html()   # HTML şablonu güncel mi kontrol et
    today_str = date.today().strftime('%Y-%m-%d')
    log.info('=' * 55)
    log.info(f'Grid Tracker calisiyor - {today_str}')
    log.info('=' * 55)

    existing = load_existing()
    trades   = read_file1(FILE1)
    overall  = read_file2(FILE2)

    if not trades and not overall:
        log.error("Her iki dosya da okunamadi."); return False

    # Excel'deki tarihi kullan (günlük dosya, tüm satırlar aynı güne ait)
    if trades:
        excel_date = trades[0]['date']
        log.info(f'Excel tarihi: {excel_date}')
    else:
        excel_date = today_str

    today_trades = [t for t in trades if t['date'] == excel_date]

    # Önceki günden devreden eşleşmemiş alışlar (FIFO kuyruğuna önce eklenir)
    # Sadece önceki bir günden üretilmişse kullan (aynı günü yeniden işliyorsak kullanma)
    op_data   = existing.get('openPositions', {})
    op_date   = op_data.get('date', '')
    prev_open = op_data.get('positions', {}) if op_date and op_date < excel_date else {}
    if prev_open:
        carried_syms = ', '.join(
            f"{sym}({sum(p['execQty'] for p in pos)} lot)"
            for sym, pos in prev_open.items()
        )
        log.info(f'Devreden acik pozisyonlar: {carried_syms}')

    profit = calc_profit(today_trades, carried_over=prev_open)

    # Overall geçmişi
    oh = existing.get('overallHistory', [])
    if overall and not any(h['date'] == excel_date for h in oh):
        oh.append({'date': excel_date, 'amount': overall,
                   'note': f'Otomatik - {datetime.now().strftime("%H:%M")}'})
        oh.sort(key=lambda x: x['date'])

    # Günlük log
    dl = existing.get('dailyLog', {})
    dl[excel_date] = {
        'date':        excel_date,
        'trades':      len(today_trades),
        'sells':       sum(1 for t in today_trades if t['type'] == 'Satış'),
        'buys':        sum(1 for t in today_trades if t['type'] == 'Alış'),
        'grossProfit': profit['totalGross'],
        'commission':  profit['totalCommission'],
        'netProfit':   profit['totalNet'],
        'overall':     overall,
        'bySymbol':    profit['bySymbol'],
        'updatedAt':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    # Tüm işlemler (geçmiş günler korunur)
    all_trades = [t for t in existing.get('allTrades', []) if t['date'] != excel_date]
    all_trades.extend(today_trades)
    all_trades.sort(key=lambda x: x['datetime'], reverse=True)

    settings = existing.get('settings', {
        'commissionRate': 1, 'gridInterval': 0, 'lotTolerance': 0,
        'costs': DEFAULT_COSTS,
    })

    # Aylık kar: ayın son BIST günündeyse otomatik hesapla ve kaydet
    monthly_kar = existing.get('monthlyKar', [])
    if overall and is_last_bist_day_of_month(date.fromisoformat(excel_date)):
        mk_profit = calc_monthly_kar(excel_date, overall, oh, existing.get('birikimTx', []))
        if mk_profit is not None:
            month_key = excel_date[:7]
            monthly_kar = [m for m in monthly_kar if m['month'] != month_key]
            monthly_kar.append({'month': month_key, 'profit': mk_profit})
            monthly_kar.sort(key=lambda x: x['month'])
            log.info(f'Aylık kar kaydedildi: {month_key} → {mk_profit:+,} ₺')

    payload = {
        'lastUpdated':    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'today':          excel_date,
        'todayOverall':   overall,
        'todayProfit':    profit,
        'allTrades':      all_trades,
        'dailyLog':       dl,
        'overallHistory': oh,
        'monthlyKar':     monthly_kar,
        # Açık pozisyon: hangi günden üretildiğini de saklıyoruz
        'openPositions':  {
            'date':      excel_date,
            'positions': profit.get('openPositions', {}),
        },
        'settings':       settings,
        # Sermaye hareketleri — kullanıcı tarafından girilir, asla sıfırlanmaz
        'birikimTx':      existing.get('birikimTx', []),
    }

    inject_into_html(payload)
    DATA_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    firebase_write(payload)

    log.info(
        f"{excel_date}: {len(today_trades)} islem | "
        f"Brut: {profit['totalGross']:+.2f} TL | "
        f"Komisyon: {profit['totalCommission']:.4f} TL | "
        f"Net: {profit['totalNet']:+.2f} TL"
    )

    # Excel dosyalarını sil (--now / dry_run modunda silme)
    if not dry_run:
        delete_excel_files()
    else:
        log.info("Test modu: Excel dosyalari silinmedi.")
    return True

# ──────────────────────────────────────────────────────────
#  EXCEL SİL  (işi biten günlük dosyaları temizle)
# ──────────────────────────────────────────────────────────
def delete_excel_files():
    for f in [FILE1, FILE2]:
        if not f.exists():
            continue
        try:
            f.unlink()
            log.info(f"Kalıcı silindi: {f.name}")
        except Exception as e:
            log.warning(f"Silinemedi ({f.name}): {e}")

# ──────────────────────────────────────────────────────────
#  GÖREV ZAMANLAYICI
# ──────────────────────────────────────────────────────────
def setup_task_scheduler():
    # Bu script evening_automation.pyw tarafından çağrılır.
    # Kendi başına görev oluşturmak çift çalışmaya neden olur.
    log.error(
        "Bu script doğrudan görev zamanlayıcıya EKLENMEZ. "
        "Görev yönetimi için evening_automation.pyw --setup kullanın."
    )

# ──────────────────────────────────────────────────────────
#  SERVİS DÖNGÜSÜ
# ──────────────────────────────────────────────────────────
def run_service():
    log.info("Grid Bot Tracker Servisi başlatıldı")
    log.info(f"Klasör : {SCRIPT_DIR}")
    log.info(f"HTML   : {HTML_FILE}")
    log.info(f"Excel 1: {FILE1}")
    log.info(f"Excel 2: {FILE2}")
    last_run = None
    while True:
        now   = datetime.now()
        today = now.date()
        if is_trading_day(today):
            rh, rm   = get_run_time(today)
            run_time = now.replace(hour=rh, minute=rm, second=0, microsecond=0)
            if last_run != today and now >= run_time:
                if run_once(): last_run = today
                time.sleep(60); continue
        time.sleep(min(seconds_until_run(), 3600))

# ──────────────────────────────────────────────────────────
#  GİRİŞ
# ──────────────────────────────────────────────────────────
if __name__ == '__main__':
    p = argparse.ArgumentParser(description='BIST Grid Bot Tracker v2')
    p.add_argument('--now',   action='store_true', help='Hemen çalıştır')
    p.add_argument('--setup', action='store_true', help='Görev zamanlayıcıya ekle')
    p.add_argument('--html',  action='store_true', help='Sadece HTML güncelle')
    args = p.parse_args()
    if args.setup:   setup_task_scheduler()
    elif args.html:  ensure_html(); log.info('HTML guncellendi.')
    elif args.now:   run_once(dry_run=False)   # Üretim: Excel dosyaları işlem sonrası silinir
    else:            run_service()

"""
price_reader.py — BIST100 anlık fiyat okuyucu
Öncelik sırası:
  1. Excel (MatriksIQ DDE) — BIST saatlerinde anlık
  2. Yahoo Finance — MatriksIQ kapalıyken
  3. Kalıcı cache (last_prices.json) — gece/hafta sonu, hiçbir kaynak yoksa son fiyat
Her geçerli fiyat kalıcı cache'e yazılır → fiyat asla boş kalmaz.
"""
import json
import time
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

_DIR        = Path(__file__).parent
EXCEL_PATH  = Path(r"C:\Users\BioCSI\CLAUDE\GridTracker\Bist100 - Anlık Fiyat.xlsx")
CACHE_PATH  = _DIR / "last_prices.json"
MAX_AGE_SECONDS = 90  # Excel dosyası bu kadar saniyeden eskiyse diskten okuma fallback

_cache = None  # {sym: {"price": float, "ts": int}}


def _load_cache() -> dict:
    global _cache
    if _cache is None:
        try:
            _cache = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            _cache = {}
    return _cache


def _save_cache(sym: str, price: float):
    """Geçerli fiyatı kalıcı cache'e yaz (son bilinen fiyat)."""
    c = _load_cache()
    c[sym.upper()] = {"price": round(price, 4), "ts": int(time.time())}
    try:
        CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


def _cache_price(sym: str) -> float | None:
    """Kalıcı cache'ten son bilinen fiyatı oku."""
    c = _load_cache()
    rec = c.get(sym.upper())
    return rec.get("price") if rec else None


def _com_init():
    """Arka plan thread'lerinde win32com erişimi için COM apartment başlat."""
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except Exception:
        pass


def _com_uninit():
    try:
        import pythoncom
        pythoncom.CoUninitialize()
    except Exception:
        pass


def _valid_price(val) -> float | None:
    """
    Geçerli fiyat mı kontrol eder.
    DDE hata kodları (negatif/büyük sayılar) ve None'ı eler.
    Makul BIST fiyat aralığı: 0.01 – 100000 TL.
    """
    try:
        p = float(val)
    except (TypeError, ValueError):
        return None
    if 0.01 <= p <= 100000:
        return round(p, 4)
    return None  # DDE hatası (-2146826265 gibi) veya saçma değer


def get_price_from_excel(symbol: str) -> float | None:
    """
    Açık Excel uygulamasından DDE verisini ANINDA okur.
    Dosya kaydını beklemez — bellek içi canlı değeri alır.
    Fallback: Excel kapalıysa diskten okur (max 90 sn eski).
    """
    sym = symbol.upper().replace(".IS", "")

    # ── Yöntem 1: Açık Excel'den canlı oku (anlık) ──────────
    _com_init()
    try:
        import win32com.client as win32
        xl = win32.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if "Anlık Fiyat" in wb.Name or "Anlik Fiyat" in wb.Name:
                ws = wb.Sheets(1)
                row = 2
                while True:
                    cell_val = ws.Cells(row, 1).Value
                    if cell_val is None:
                        break
                    if str(cell_val).upper() == sym:
                        price = ws.Cells(row, 3).Value  # C = SON
                        p = _valid_price(price)
                        if p:
                            return p
                        break  # geçersiz/DDE hatası → fallback'e geç
                    row += 1
    except:
        pass

    # ── Yöntem 2: Diskten oku (Excel kapalıysa fallback) ────
    try:
        if not EXCEL_PATH.exists():
            return None
        age = time.time() - EXCEL_PATH.stat().st_mtime
        if age > MAX_AGE_SECONDS:
            return None
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).upper() == sym:
                p = _valid_price(row[2])
                wb.close()
                return p
        wb.close()
    except:
        pass

    return None


def get_price_from_yahoo(symbol: str) -> float | None:
    """Yahoo Finance fallback."""
    try:
        import yfinance as yf
        t = yf.Ticker(f"{symbol}.IS")
        hist = t.history(period="2d")
        if not hist.empty:
            return float(hist["Close"].iloc[-1])
    except:
        pass
    return None


def get_price(symbol: str) -> tuple[float | None, str]:
    """
    Fiyat döndürür. 3 katmanlı, asla boş kalmaz.
    Returns: (fiyat, kaynak) — kaynak: 'excel' | 'yahoo' | 'cache'
    """
    sym = symbol.upper().replace(".IS", "")

    # 1. Excel (MatriksIQ DDE) — BIST saatlerinde anlık
    price = get_price_from_excel(sym)
    if price and price > 0:
        _save_cache(sym, price)   # son fiyatı kalıcı yaz
        return price, "excel"

    # 2. Yahoo Finance — MatriksIQ kapalıyken
    price = get_price_from_yahoo(sym)
    if price and price > 0:
        _save_cache(sym, price)
        return price, "yahoo"

    # 3. Kalıcı cache — gece/hafta sonu, hiçbir canlı kaynak yok
    price = _cache_price(sym)
    if price and price > 0:
        return price, "cache"

    return None, "yok"


def get_all_prices() -> dict[str, float]:
    """
    Tüm BIST100 fiyatlarını döner.
    Geçerli değerler kalıcı cache'e yazılır; boş gelenler cache'ten doldurulur.
    Böylece BIST kapalıyken de son fiyatlar korunur.
    """
    prices = {}

    # ── Açık Excel'den canlı oku ─────────────────────────────
    _com_init()
    try:
        import win32com.client as win32
        xl = win32.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if "Anlık Fiyat" in wb.Name or "Anlik Fiyat" in wb.Name:
                ws = wb.Sheets(1)
                row = 2
                while True:
                    sym = ws.Cells(row, 1).Value
                    if sym is None:
                        break
                    p = _valid_price(ws.Cells(row, 3).Value)  # DDE hata filtresi
                    if p:
                        prices[str(sym).upper()] = p
                    row += 1
                break
    except:
        pass

    # ── Excel boş döndüyse diskten dene ──────────────────────
    if not prices:
        try:
            if EXCEL_PATH.exists():
                age = time.time() - EXCEL_PATH.stat().st_mtime
                if age <= MAX_AGE_SECONDS:
                    import openpyxl
                    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
                    ws = wb.active
                    for r in ws.iter_rows(min_row=2, values_only=True):
                        if r[0]:
                            p = _valid_price(r[2])
                            if p:
                                prices[str(r[0]).upper()] = p
                    wb.close()
        except:
            pass

    # ── Geçerli fiyatları kalıcı cache'e yaz ─────────────────
    if prices:
        c = _load_cache()
        for sym, p in prices.items():
            c[sym] = {"price": p, "ts": int(time.time())}
        try:
            CACHE_PATH.write_text(json.dumps(c, ensure_ascii=False), encoding="utf-8")
        except:
            pass

    # ── Eksikleri/boşları kalıcı cache'ten doldur ────────────
    # (BIST kapalı → Excel hata → cache'teki son fiyatlar kullanılır)
    c = _load_cache()
    for sym, rec in c.items():
        if sym not in prices and rec.get("price"):
            prices[sym] = rec["price"]

    return prices


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    sym = sys.argv[1].upper() if len(sys.argv) > 1 else "SASA"
    price, src = get_price(sym)
    print(f"{sym}: {price} TL  (kaynak: {src})")
    age = time.time() - EXCEL_PATH.stat().st_mtime if EXCEL_PATH.exists() else -1
    print(f"Excel yaşı: {age:.0f} sn")
